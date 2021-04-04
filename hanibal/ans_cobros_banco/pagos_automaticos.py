# -*- coding: utf-8 -*-
from openerp.osv import osv,fields
from openerp import models, fields, api, _
from datetime import datetime,timedelta,date
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import io
import logging
import json
import base64
import os.path
import csv
import os
import binascii
import struct
_logger = logging.getLogger(__name__)
DATE_FORMAT = '%Y-%m-%d'

class PagosAutomaticosFacturas(models.Model):
    _name="pagos.automaticos.facturas"

    cabecera_id =fields.Many2one('pagos.automaticos.detalle',string="Relacion")
    estado = fields.Selection(related='cabecera_id.estado',string="Estado")
    estudiante_id = fields.Many2one('res.partner',string="Estudiante")
    factura_id = fields.Many2one('account.invoice',string="Factura",domain="[('state','=','open'),('alumno_id','=',estudiante_id),('residual','!=',0.0)]")
    fecha_factura = fields.Date(related='factura_id.date_invoice',string="Fecha")
    saldos = fields.Float(related='factura_id.residual',string="Saldo")
    total = fields.Float(related='factura_id.amount_total',string="Total")

    # @api.constrains('factura_id')
    # def validar_valor_contrans(self):
    #     obj_cabecera=self.env['pagos.automaticos.detalle'].search([('id','=',self.cabecera_id.id)])
    #     contador=0.0
    #     for l in obj_cabecera:
    #         for m in l.pagos_facturas:
    #             if self.factura_id.id == m.factura_id.id:
    #                 comentario=("Factura ya ingresada.\n ")
    #                 raise osv.except_osv(('Alerta'),comentario)

    # @api.multi
    # def validar_valor(self):
    #     obj_cabecera=self.env['pagos.automaticos.detalle'].search([('id','=',self.cabecera_id.id)])
    #     contador=0.0
    #     for l in obj_cabecera:
    #         for m in l.pagos_facturas:
    #             if self.factura_id.id == m.factura_id.id:
    #                 comentario=("Factura ya ingresada.\n ")
    #                 raise osv.except_osv(('Alerta'),comentario)


class PagosAutomaticosDetalle(models.Model):
    _name="pagos.automaticos.detalle"

    pagos_id =fields.Many2one('pagos.automaticos',string="Relacion")
    pagos_facturas=fields.One2many('pagos.automaticos.facturas','cabecera_id',string="Relacion")
    estado = fields.Selection(related='pagos_id.estado',string="Estado")
    num_orden = fields.Char(string="Orden")
    fecha_creacion= fields.Date('Fecha')
    num_contrato = fields.Char(related='estudiante_id.codigo_alumno',string="Contrato",store=True)
    servicios = fields.Char(string="Servicio")
    concepto = fields.Char(string="Concepto")
    identificacion = fields.Char(related='estudiante_id.vat',string="Identificacion",store=True)
    estudiante_id = fields.Many2one('res.partner',string="Estudiante")
    nom_benef = fields.Char(related='estudiante_id.name',string="Beneficiario",store=True)
    forma_cobro = fields.Char(string="Forma de cobro")
    cuenta_banco=fields.Char('Cuenta',size=20)
    moneda = fields.Char(string="Moneda")
    valor = fields.Float(string="Valor")
    saldo = fields.Float(string="Saldo")
    est_detale = fields.Char(string="Detalle")
    cod_error = fields.Char(string="Codigo Error")
    desc_error = fields.Char(string="Desc Error")
    texto1 = fields.Char(string="TEXTO1")
    texto2 = fields.Char(string="TEXTO2")
    #factura = fields.Char(string="Factura")
    factura_id = fields.Many2many('account.invoice',string="Factura",domain="[('state','=','open'),('residual','!=',0.0)]")
    error_sistema = fields.Char(string="Error Sistema")
    descripcion =fields.Text(string="Descripcion")
    error = fields.Char(string="Error",default='N')

    @api.constrains('pagos_facturas')
    def validar_facturas(self):
        contador=0.0
        list_fact=[]
        for l in self.pagos_facturas:
            list_fact.append(l.factura_id.id)
            contador=contador+l.total
        repetido = []
        unico = []
        for x in list_fact:
            if x not in unico:
                unico.append(x)
            else:
                if x not in repetido:
                    repetido.append(x)

        if len(repetido)!=0:
            comentario=("Factura ya ingresada.\n ")
            raise osv.except_osv(('Alerta'),comentario)


        if contador!=0:
            if self.valor!=contador:
                if self.valor<contador:
                    comentario=("Contrato: {0} - Valor Cobro: {1} - Valor Facturas: {2} \n El monto de las facturas es mayor que el valor a pagar.\n ").format(self.num_contrato,self.valor,contador)
                    raise osv.except_osv(('Alerta1'),comentario)
                else:
                    comentario=("Contrato: {0} - Valor Cobro: {1} - Valor Facturas: {2} \n El monto de las facturas es menor que el valor a pagar.\n ").format(self.num_contrato,self.valor,contador)
                    raise osv.except_osv(('Alerta1'),comentario)
        else:
            pass

    @api.multi
    def llenar_descripcion(self):
        if len(self.factura_id)!=0:
            for f in self.factura_id:
                if len(f.invoice_line)!=0:
                    descripcion=''
                    for d in f.invoice_line:
                        descripcion=descripcion+d.name+' '
                    self.descripcion=descripcion
                    return {
                        'value': {
                            'descripcion': descripcion,
                            }
                        }

        else:
            descripcion=''
            self.descripcion=descripcion
            return {
                'value': {
                    'descripcion': '',
                    }
                }

    @api.multi
    def guardar_facturas(self):
        lista=[]
        contador=0.0
        for l in self.pagos_facturas:
            lista.append(l.factura_id.id)
            contador=contador+l.total
        if len(self.pagos_facturas)!=0:
            self.factura_id=None
            self.factura_id=lista
            self.write({
                'factura_id':lista
                })
        else:
            self.factura_id=None
            self.write({
                'factura_id':[]
                })



    @api.multi
    def action_from(self):
        lista=[]
        contador=0.0
        for l in self.pagos_facturas:
            lista.append(l.factura_id.id)
        if len(self.pagos_facturas)!=0:
            self.factura_id=None
            self.factura_id=lista
        else:
            self.factura_id=None
        viewid = self.env.ref('ans_cobros_banco.view_cobros_factura_cabezera_form').id
        context = self._context.copy()
        obj_descuenta_detalle=self.env['pagos.automaticos.detalle'].search([('id','=',self.id)],limit=1)
        return {   
                'name':'Detalle de Facturas',
                'res_model': 'pagos.automaticos.detalle',
                'view_type': 'form',
                'views' : [(viewid,'form')],
                'type':'ir.actions.act_window',
                'nodestroy': True,
                'res_id': obj_descuenta_detalle.id,
                'target':'new',
                'context':context,
                }



class PagosAutomaticos(models.Model):
    _name="pagos.automaticos"
    _rec_name = "display_name"

    usuario_id=fields.Many2one('res.users',string="Usuario", default=lambda self:self.env.user ,readonly=True)
    codigo=fields.Char(string='Codigo')
    banco_id=fields.Many2one('res.partner.bank','Banco',copy=False, index=True)
    cuenta_banco=fields.Char(related='banco_id.acc_number',string='Cuenta Bancaria',size=20)
    fecha_creacion= fields.Datetime('Fecha Creacion', readonly=True, copy=False,select=True,)
    archivo = fields.Binary(string='Archivo')
    filename = fields.Char(string="Nombre")
    pagos_line=fields.One2many('pagos.automaticos.detalle','pagos_id',string="Relacion")
    estado = fields.Selection( (('0','Borrador'),
                               ('1','Validado')) , 'Estado', required=False,default='0')
    display_name =fields.Char(string="Nombre a mostrar")
    journal_id=fields.Many2one('account.journal',string='Metodo de Pago')
    notas =fields.Text(string="Notas")
    account_journal_caja_id = fields.Many2one('account.journal',string="Caja",default=lambda self: self._default_route_ids())
    exigir_doc_banco = fields.Boolean(related='journal_id.exigir_doc_banco',string="Exigir datos bancarios en pago")
    usuario_id = fields.Many2one('res.users',string="Usuario", default=lambda self:self.env.user ,readonly=True)
    fecha_cobro = fields.Date(string='Fecha Cobro')
    active = fields.Boolean(string="Activo",default=False)

    @api.model
    def _default_route_ids(self):
        obj_datos=self.env['account.journal'].search([('usuarios_ids','in',self._uid),('caja','=',True)])
        if len(obj_datos)==1:
            return obj_datos.id
        else:
            return False

    @api.constrains('codigo','banco_id','cuenta_banco')
    def guardar_display_name(self):
        if self.codigo==False:
            self.display_name=str(str('Cobro Automatico'))
        else:
            self.display_name=str(str('Cobro Automatico')+'/'+str(self.codigo))

    _defaults = {
        'fecha_creacion': fields.datetime.now(),
    }

    @api.multi
    def cargar_archivo(self):
        self.validar_xlsx(self.filename,self.archivo)
        

    @api.multi
    def validar_xlsx(self, filename, file):
        tmp = filename.split('.')
        ext = tmp[len(tmp)-1]
        if ext=='xlsx':
            self.cargar_xlsx(file)
        else:
            raise osv.except_osv(('Alerta'),("El tipo de archivo debe ser xlsx."))

    def cambiar_formato(self,fecha):
        datetime_object = datetime.strptime(fecha, DATE_FORMAT)
        fecha_dato= datetime_object.date()
        return fecha_dato

    @api.multi
    def cargar_xlsx(self,file):
        #ssssssssssssstry:
            decoded_data = base64.b64decode(file)
            xls_file = io.BytesIO(decoded_data)
            wb=load_workbook(filename=xls_file,read_only=True)
            count=0
            data2=[]
            ws = wb.active
            for i in range(1,ws.max_row+1):
                data=[]
                for j in range(1,ws.max_column+1):

                    cell_obj=ws.cell(row=i,column=j)
                    data.append(cell_obj.value)
            
                data2.append(data)
            l=[]
            orden=0
            self.env.cr.execute("delete from pagos_automaticos_detalle where pagos_id={0}".format(self.id))
            for l in data2:
                NUM_ORDEN=''
                FECHA=''
                NUM_CONTRATO=''
                SERVICIO=''
                CONCEPTO=''
                IDENTFICACION=''
                NOM_BENEF=''
                FORMA_COBRO=''
                CUENTA=''
                MONEDA=''
                VALOR=''
                SALDO=''
                EST_DETALLE=''
                COD_ERROR=''
                DESC_ERROR=''
                TEXTO1=''
                TEXTO2=''
                if orden>=1:
                    NUM_ORDEN=str(l[0])
                    FECHA=str(l[1])
                    NUM_CONTRATO=str(l[2])
                    SERVICIO=str(l[3])
                    CONCEPTO=str(l[4])
                    IDENTFICACION=str(l[5])
                    NOM_BENEF=str(l[6])
                    FORMA_COBRO=str(l[7])
                    CUENTA=str(l[8])
                    MONEDA=str(l[9])
                    VALOR=str(l[10])
                    SALDO=str(l[11])
                    EST_DETALLE=str(l[12])
                    COD_ERROR=str(l[13])
                    DESC_ERROR=str(l[14])
                    TEXTO1=str(l[15])
                    TEXTO2=str(l[16])
                    anio=FECHA[0:4]
                    mes=FECHA[4:6]
                    dia=FECHA[6:8]
                    fecha_registro=str(anio+'-'+mes+'-'+dia)
                    NUM_ORDEN=NUM_ORDEN.split('"')
                    NUM_CONTRATO=NUM_CONTRATO.split('"')
                    if len(NUM_CONTRATO[1])!=0:
                        NUM_CONTRATO=NUM_CONTRATO[1].replace(' ', '')
                    CUENTA=CUENTA.split('"')
                    if len(CUENTA[1])!=0:
                        CUENTA=CUENTA[1].replace(' ', '')
                    VALOR=VALOR.split('"')
                    if len(VALOR[1])!=0:
                        VALOR=VALOR[1].replace(' ', '')
                    SALDO=SALDO.split('"')
                    if len(SALDO[1])!=0:
                        SALDO=SALDO[1].replace(' ', '')
                    CONCEPTO_FECHA=CONCEPTO.split(' ')
                    CONCEPTO_FECHA=CONCEPTO_FECHA[1].split('/')
                    fecha_factura=str(CONCEPTO_FECHA[2]+'-'+CONCEPTO_FECHA[1]+'-'+CONCEPTO_FECHA[0])
                    obj_datos=self.env['res.partner'].search([('codigo_alumno','like',str(NUM_CONTRATO))])
                    if self.codigo==False:
                        obj_orden=self.env['pagos.automaticos'].search([('codigo','=',NUM_ORDEN[1])])
                        if len(obj_orden)>1:
                            raise osv.except_osv(('Alerta'),("El archivo cuenta con un codigo ya cargado al sistema."))
                        else:
                            self.codigo=NUM_ORDEN[1]
                    if obj_datos:
                        if len(obj_datos)>1:
                            for l in obj_datos:
                                obj_factura=self.env['account.invoice'].search([('alumno_id','=',l.id),('residual','=',VALOR),('date_invoice','=',fecha_factura)])
                                if obj_factura.id:
                                    lista_fact=[obj_factura.id]
                                else:
                                    lista_fact=[]
                                descripcion=''
                                for fact in obj_factura.invoice_line:
                                    descripcion=descripcion+fact.name+' '
                                obj_detalle=self.env['pagos.automaticos.detalle'].create({
                                    'num_orden':NUM_ORDEN[1],
                                    'fecha_creacion':fecha_registro,
                                    'num_contrato':l.codigo_alumno,
                                    'servicios':SERVICIO,
                                    'concepto':CONCEPTO,
                                    'identificacion':l.vat,
                                    'nom_benef':l.name,
                                    'estudiante_id':l.id,
                                    'forma_cobro':FORMA_COBRO,
                                    'cuenta_banco':CUENTA,
                                    'moneda':MONEDA,
                                    'valor':VALOR,
                                    'saldo':SALDO,
                                    'est_detale':EST_DETALLE,
                                    'cod_error':COD_ERROR,
                                    'desc_error':DESC_ERROR,
                                    'descripcion':descripcion,
                                    #'texto1':'',
                                    #'texto2':'',
                                    'pagos_id':self.id,
                                    #'factura_id':lista_fact
                                    })
                                if len(lista_fact)!=0:                                
                                    obj_detalle.factura_id=lista_fact
                                else:
                                    obj_detalle.factura_id=False
                                self.active=True
                            self.validar_facturas()
                        else:
                            obj_datos=self.env['res.partner'].search([('codigo_alumno','like',str(NUM_CONTRATO))])
                            descripcion=''
                            obj_factura=self.env['account.invoice'].search([('alumno_id','=',obj_datos.id),('residual','=',VALOR),('date_invoice','=',fecha_factura)])
                            
                            if obj_factura.id:
                                lista_fact=[obj_factura.id]
                            else:
                                lista_fact=[]
                            for fact in obj_factura.invoice_line:
                                descripcion=descripcion+fact.name+' '
                            obj_detalle=self.env['pagos.automaticos.detalle'].create({
                                    'num_orden':NUM_ORDEN[1],
                                    'fecha_creacion':fecha_registro,
                                    'num_contrato':obj_datos.codigo_alumno,
                                    'servicios':SERVICIO,
                                    'concepto':CONCEPTO,
                                    'identificacion':obj_datos.vat,
                                    'nom_benef':obj_datos.name,
                                    'estudiante_id':obj_datos.id,
                                    'forma_cobro':FORMA_COBRO,
                                    'cuenta_banco':CUENTA,
                                    'moneda':MONEDA,
                                    'valor':VALOR,
                                    'saldo':SALDO,
                                    'est_detale':EST_DETALLE,
                                    'cod_error':COD_ERROR,
                                    'desc_error':DESC_ERROR,
                                    'descripcion':descripcion,
                                    'pagos_id':self.id,
                                    })
                            if len(lista_fact)!=0:                        
                                obj_detalle.factura_id=lista_fact
                            else:
                                obj_detalle.factura_id=False
                            self.active=True
                            self.validar_facturas()
                    else:
                        descripcion=""
                        obj_detalle=self.env['pagos.automaticos.detalle'].create({
                                    'num_orden':NUM_ORDEN[1],
                                    'fecha_creacion':fecha_registro,
                                    'num_contrato':str(NUM_CONTRATO),
                                    'servicios':SERVICIO,
                                    'concepto':CONCEPTO,
                                    'identificacion':str(IDENTFICACION),
                                    'nom_benef':str(NOM_BENEF),
                                    'forma_cobro':FORMA_COBRO,
                                    'cuenta_banco':CUENTA,
                                    'moneda':MONEDA,
                                    'valor':VALOR,
                                    'saldo':SALDO,
                                    'est_detale':EST_DETALLE,
                                    'cod_error':'Error del Número de Contrato',
                                    'desc_error':DESC_ERROR,
                                    'descripcion':descripcion,
                                    'error':'S',
                                    'pagos_id':self.id,
                                    })
                        self.env.cr.execute("update pagos_automaticos_detalle set num_contrato='{0}', nom_benef='{1}', identificacion='{2}' where pagos_id={3} and id={4}".format(str(NUM_CONTRATO),str(NOM_BENEF),str(IDENTFICACION),self.id,obj_detalle.id))
                orden=1

            
            self.estado='0'
        # except Exception as e:
        #     comentario=("Por favor revisar el formato del archivo y los datos sean correctos.\n Fechas \n Concepto \n Saldo \n Valor \n Celdas vacias ")
        #     raise osv.except_osv(('Alerta'),comentario)


    @api.constrains('pagos_line')
    def validar_facturas(self):
        for l in self.pagos_line:
            cont=0.0
            #self.env.cr.execute("delete from pagos_automaticos_facturas where cabecera_id={0}".format(l.id))
            if len(l.factura_id)!=0:
                for d in l.factura_id:
                    cont=cont+d.residual
                    if len(d.invoice_line)!=0:
                        descripcion=''
                        for a in d.invoice_line:
                            descripcion=descripcion+a.name+' '
                l.descripcion=descripcion
                if l.valor!=cont:
                    if l.valor<cont:
                        comentario=("Contrato: {0} - Valor Cobro: {1} - Valor Facturas: {2} \n El monto de las facturas es mayor que el valor a pagar.\n ").format(l.num_contrato,l.valor,cont)
                        raise osv.except_osv(('Alerta2'),comentario)
                    else:
                        comentario=("Contrato: {0} - Valor Cobro: {1} - Valor Facturas: {2} \n El monto de las facturas es menor que el valor a pagar.\n ").format(l.num_contrato,l.valor,cont)
                        raise osv.except_osv(('Alerta2'),comentario)
                else:
                    for f in l.factura_id:
                        obj_factura_d = self.env['pagos.automaticos.facturas']
                        obj_valida = self.env['pagos.automaticos.facturas'].search([('cabecera_id','=',l.id),('factura_id','=',f.id)])
                        if len(obj_valida)==0:
                            objeto = obj_factura_d.create({
                                'cabecera_id':l.id,
                                'estudiante_id':l.estudiante_id.id,
                                'factura_id':f.id,
                                })
            else:
                l.descripcion=''
                self.env.cr.execute("delete from pagos_automaticos_facturas where cabecera_id={0}".format(l.id))
                    
    @api.multi
    def validar_cobros(self):
        # try:
            for l in self.pagos_line:
                if len(l.factura_id)==0:
                    #l.error_sistema='No cuenta con factura para realizar el pago'
                    #l.error='S'
                    obj_cobro=self.env['account.voucher']
                    obj_cobro_linea=self.env['account.voucher.line']
                    #obj_factura=self.env['account.invoice'].search([('id','=',f.id)])
                    #obj_factura.invoice_pay_customer()
                    #obj_factura.button_proforma_voucher()
                    print(l.estudiante_id.id,'l.estudiante_id')
                    print(l.moneda,'l.moneda')
                    obj_alumno=self.env['res.partner'].search([('id','=',int(l.estudiante_id.id))])
                    obj_moneda=self.env['res.currency'].search([('name','like',l.moneda)])
                    print(obj_moneda.id,'obj_moneda')
                    print(obj_alumno.id,'obj_alumno')
                    print(obj_alumno.parent_id.id,'partner_id')
                    obj_dato_c=obj_cobro.create({
                        'date_due':self.fecha_cobro,
                        'escolar':True,
                        'alumno_id':obj_alumno.id,
                        'partner_id':obj_alumno.parent_id.id,
                        'amount':float(l.valor),
                        'comment':'Write-Off',
                        'account_journal_caja_id':self.account_journal_caja_id.id,
                        'journal_id': self.journal_id.id,
                        'payment_rate_currency_id':obj_moneda.id,
                        'account_id': self.journal_id.default_debit_account_id.id,
                        'company_id': obj_alumno.company_id.id,
                        'payment_rate':1.000000,
                        #'payment_option':'without_writeoff',
                        #'currency_id': obj_factura.currency_id.id,
                        #'reference': obj_factura.name,
                        #'invoice_type': obj_factura.type,
                        #'invoice_id': obj_factura.id,
                        #'default_type': inv.type in ('out_invoice','out_refund') and 'receipt' or 'payment',
                        'type': 'receipt',
                        })
                    #obj_move_p = self.env['account.move.line'].search([('account_id','=',obj_factura.account_id.id),('ref','=',str(obj_factura.number))])
                    # obj_data_c_d = obj_cobro_linea.create({
                    #     'reconcile':True,
                    #     'amount_unreconciled':l.valor,
                    #     'company_id':self.usuario_id.company_id.id,
                    #     'type':'cr',
                    #     #'account_id':obj_factura.account_id.id,
                    #     'voucher_id':obj_dato_c.id,
                    #     'amount_original':l.valor,
                    #     'name':'',
                    #     #'move_line_id':obj_move_p.id,
                    #     'amount':l.valor,
                    #     })

                    if obj_dato_c:# and obj_data_c_d:
                       obj_dato_c.proforma_voucher()
                       obj_dato_c.button_proforma_voucher()
                    #comentario=("Contrato: {0} - Valor Cobro: {1} \n Debe seleccionar factura a pagar.\n ").format(l.num_contrato,l.valor)
                    #raise osv.except_osv(('Alerta'),comentario)
                else:
                    for f in l.factura_id:
                        obj_cobro=self.env['account.voucher']
                        obj_cobro_linea=self.env['account.voucher.line']
                        obj_factura=self.env['account.invoice'].search([('id','=',f.id)])
                        obj_factura.invoice_pay_customer()
                        #obj_factura.button_proforma_voucher()
                        obj_dato_c=obj_cobro.create({
                            'date_due':self.fecha_cobro,
                            'escolar':True,
                            'alumno_id':obj_factura.alumno_id.id,
                            'partner_id':obj_factura.partner_id.id,
                            'amount':obj_factura.residual,
                            'account_journal_caja_id':self.account_journal_caja_id.id,
                            'currency_id': obj_factura.currency_id.id,
                            'reference': obj_factura.name,
                            'invoice_type': obj_factura.type,
                            'invoice_id': obj_factura.id,
                            'journal_id': self.journal_id.id,
                            'account_id': self.journal_id.default_debit_account_id.id,
                            #'default_type': inv.type in ('out_invoice','out_refund') and 'receipt' or 'payment',
                            'type': obj_factura.type in ('out_invoice','out_refund') and 'receipt' or 'payment',
                            })
                        obj_move_p = self.env['account.move.line'].search([('account_id','=',obj_factura.account_id.id),('ref','=',str(obj_factura.number))])
                        obj_data_c_d = obj_cobro_linea.create({
                            'reconcile':True,
                            'amount_unreconciled':obj_factura.residual,
                            'company_id':self.usuario_id.company_id.id,
                            'type':'cr',
                            'account_id':obj_factura.account_id.id,
                            'voucher_id':obj_dato_c.id,
                            'amount_original':obj_factura.residual,
                            'name':'',
                            'move_line_id':obj_move_p.id,
                            'amount':obj_factura.residual,
                            })

                        if obj_dato_c and obj_data_c_d:
                            obj_dato_c.proforma_voucher()
                            obj_dato_c.button_proforma_voucher()

            self.estado='1'
            viewid = self.env.ref("ans_cobros_banco.cerrar_ventana_cobros_falla").id
            return {
            'name':'Cobros Generados con Éxito!',
            'view_type':'form',
            'views' : [(viewid,'form')],
            'res_model':'close.window.purchase',
            'type':'ir.actions.act_window',
            'target':'new',
            }
        # except Exception as e:
        #     comentario=("Fallo al generar los cobros.")
        #     raise osv.except_osv(('Alerta'),comentario)



class account_invoice_cobrod(models.Model):
    _inherit= "account.invoice"
    _rec_name = "numerofac"



    @api.multi
    def name_get(self):
        if self._context is None:
            self._context = {}
        res = []
        if self._context.get('nombre_especial_para_mostrar', False):
            for partner in self:
                for inv in self:
                    res.append((inv.id, "%s, Saldo: %s, Total: %s" % (inv.number or '', inv.residual or '0.0', inv.amount_total or '0.0')))
        else:
            TYPES = {
            'out_invoice': _('Invoice'),
            'in_invoice': _('Supplier Invoice'),
            'out_refund': _('Refund'),
            'in_refund': _('Supplier Refund'),
            }
            res = []
            for inv in self:
                res.append((inv.id, "%s %s" % (inv.number or TYPES[inv.type], inv.name or '')))
        return res
