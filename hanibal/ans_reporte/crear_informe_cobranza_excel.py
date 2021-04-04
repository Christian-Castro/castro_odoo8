# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from copy import deepcopy
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, date, timedelta
import time
import locale
from openpyxl.worksheet.pagebreak import Break
global root

def crear_wb_informe():
    wb = openpyxl.Workbook()
    return wb


def unicodeText(text):
    try:
        text = unicodedata.unicode(text, 'utf-8')
    except TypeError:
        return text

def crea_hoja_info(wb, title, flag):
    if(flag == 0):
        sheet = wb.active
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
    if(flag == 1):
        sheet = wb.create_sheet()
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
    sheet.title = title
    return sheet

def crea_hoja_info_pdf(wb, title, flag, tutor=False):
    sheet = wb.active
    if(flag == 0):
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        if tutor:
            sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT #vertical
        else:
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE #horizontal
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = False
        sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet



def border_tabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

    colfin=colfin+1
    filfin=filfin+2

    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    for i in range(fil, filfin-1):
        for j in range(col, colfin):
            sheet.cell(row=i, column=j).border = border_cell


def columnas_filas(sheet, flag, celda, value):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value



def poner_border(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    sheet.cell(row=fil, column=col).border = border_cell

def aumentar_letra(letra):
    letra = ord(letra) + 1 #obtengo el numero ascii
    letra = chr(letra) #obtengo el caracter ascii
    return letra

def Informe_financiero(sheet, dic, lista_alumnos, cant_alumno, v_saldo, v_total, tutor=False):
    columnas_filas(sheet, 0, 'A', 5.00)
    columnas_filas(sheet, 0, 'B', 8.00)
    columnas_filas(sheet, 0, 'C', 35.00)     
    columnas_filas(sheet, 0, 'D', 10.00)
    columnas_filas(sheet, 0, 'E', 10.00)
    columnas_filas(sheet, 0, 'F', 10.00)
    columnas_filas(sheet, 0, 'G', 8.00)
    if not tutor:
        if v_saldo and v_total:
            columnas_filas(sheet, 0, 'L', 29.00)
        elif v_saldo or v_total:
            columnas_filas(sheet, 0, 'K', 29.00)
        else:
            columnas_filas(sheet, 0, 'J', 29.00)
    else:
        columnas_filas(sheet, 0, 'F', 35.00)

    fuente = Font(bold=False, size=7, name='arial')
    fuente3 = Font(bold=True, size=10, name='arial')
    fuente2 = Font(bold=True, size=7, name='arial')
    formato_numero = '#,##0.00'

    fila = 3

    if not tutor:
        if v_saldo and v_total:
            sheet.merge_cells('A2:L2')
        elif v_saldo or v_total:
            sheet.merge_cells('A2:K2')
        else:
            sheet.merge_cells('A2:J2')
    else:
        sheet.merge_cells('A2:F2')

    sheet['A2'].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente3
    if dic['seccion_id']==False:
        sheet['A2']= 'COBRANZAS'+' '+str(dic['jornada_id'])
    else:
        sheet['A2']= 'COBRANZAS'+' '+str(dic['seccion_id'].encode('utf-8'))+' '+str(dic['jornada_id'])

    if not tutor:
        if v_saldo and v_total:
            sheet.merge_cells('A3:L3')
        elif v_saldo or v_total:
            sheet.merge_cells('A3:K3')
        else:
            sheet.merge_cells('A3:J3')
    else:
        sheet.merge_cells('A3:F3')

    sheet['A3'].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
    sheet['A3'].font = fuente2
    sheet['A3'] = str(dic['fecha_corte'])

    fila=3
    total_general = 0.0
    saldo_general = 0.0
    total_general_dias_mora = 0
    total_general_pagos = 0
    total_general_cheques_postfechados = 0
    for recorrer in lista_alumnos:
        columnas_filas(sheet, 1, str(fila+2), 10.00)
        columnas_filas(sheet, 1, str(fila+3), 10.00)
        sheet.merge_cells('A'+str(fila+2)+':B'+str(fila+2))
        sheet['A'+str(fila+2)].alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+2)].font = fuente2
        sheet['A'+str(fila+2)] = 'Fecha de corte:'

        sheet['C'+str(fila+2)].alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila+2)].font = fuente
        sheet['C'+str(fila+2)] = str(dic['fecha_corte'])

        sheet['A'+str(fila+3)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['A'+str(fila+3)].font = fuente
        sheet['A'+str(fila+3)] = 'Curso'

        sheet.merge_cells('B'+str(fila+3)+':C'+str(fila+3))
        sheet['B'+str(fila+3)].alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+3)].font = fuente
        sheet['B'+str(fila+3)] = (recorrer['curso'])

        sheet['D'+str(fila+3)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['D'+str(fila+3)].font = fuente
        sheet['D'+str(fila+3)] = 'Paralelo: '+str(recorrer['paralelo'].encode('utf-8'))

        if dic['seccion_id']==False:
            sheet['E'+str(fila+3)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet['E'+str(fila+3)].font = fuente
            sheet['E'+str(fila+3)] = 'Seccion: '+str(recorrer['seccion'].encode('utf-8'))

        sheet['A'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['A'+str(fila+5)].font = fuente2
        sheet['A'+str(fila+5)] = 'Tipo'
        
        sheet['B'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['B'+str(fila+5)].font = fuente2
        sheet['B'+str(fila+5)] = 'Código Alumno'

        sheet['C'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['C'+str(fila+5)].font =fuente2
        sheet['C'+str(fila+5)] = 'Alumno'

        sheet['D'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['D'+str(fila+5)].font = fuente2
        sheet['D'+str(fila+5)] = 'Número'

        sheet['E'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['E'+str(fila+5)].font = fuente2
        sheet['E'+str(fila+5)] = 'Emision'

        if not tutor:
            sheet['F'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet['F'+str(fila+5)].font = fuente2
            sheet['F'+str(fila+5)] = 'Vencimiento'

            sheet['G'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet['G'+str(fila+5)].font = fuente2
            sheet['G'+str(fila+5)] = 'Días Mora'

            letra = "H"
            if v_total:
                sheet['H'+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
                sheet['H'+str(fila+5)].font = fuente2
                sheet['H'+str(fila+5)] = 'Valor'
                letra = "I"
            
            sheet[letra+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet[letra+str(fila+5)].font = fuente2
            sheet[letra+str(fila+5)] = 'Pagos'

            letra = aumentar_letra(letra)
            sheet[letra+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet[letra+str(fila+5)].font = fuente2
            sheet[letra+str(fila+5)] = 'Cheques Postfechados'

            if v_saldo:
                letra = aumentar_letra(letra)
                sheet[letra+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
                sheet[letra+str(fila+5)].font = fuente2
                sheet[letra+str(fila+5)] = 'Saldo Actual'

        if tutor:
            letra = "F"
        else:
            letra = aumentar_letra(letra)
        sheet[letra+str(fila+5)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet[letra+str(fila+5)].font = fuente2
        sheet[letra+str(fila+5)] = 'Comentario'

        fila += 6
        saldo = 0
        total = 0
        total_dias_mora = 0
        total_pagos = 0
        total_cheques_postfechados = 0
        for det in recorrer['detalle']:
            columnas_filas(sheet, 1, str(fila), 10.00)
            sheet['A'+str(fila)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet['A'+str(fila)].font = fuente
            sheet['A'+str(fila)] = det['tipo']

            sheet['B'+str(fila)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet['B'+str(fila)].font = fuente
            sheet['B'+str(fila)] = det['codigo_alumno'] if det['codigo_alumno'] else ''

            sheet['C'+str(fila)].alignment = Alignment(wrapText=True,horizontal='justify', vertical='top')
            sheet['C'+str(fila)].font = fuente
            sheet['C'+str(fila)] = det['alumno'] if det['alumno'] else ''

            sheet['D'+str(fila)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet['D'+str(fila)].font = fuente
            sheet['D'+str(fila)] = det['numero'] if det['numero'] else ''

            sheet['E'+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
            sheet['E'+str(fila)].font = fuente
            sheet['E'+str(fila)] = det['emision'] if det['emision'] else ''

            if not tutor:
                sheet['F'+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
                sheet['F'+str(fila)].font = fuente
                sheet['F'+str(fila)] = det['vencimiento']

                sheet['G'+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
                sheet['G'+str(fila)].font = fuente
                sheet['G'+str(fila)] = det['dias_mora']
                total_dias_mora += det['dias_mora']

                letra = "H"
                if v_total:
                    sheet['H'+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
                    sheet['H'+str(fila)].font = fuente
                    sheet['H'+str(fila)].number_format = formato_numero
                    sheet['H'+str(fila)] = det["total"]
                    letra = "I"

                sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
                sheet[letra+str(fila)].font = fuente
                sheet[letra+str(fila)].number_format = formato_numero
                sheet[letra+str(fila)] = det['pagos']
                total_pagos += det['pagos']

                letra = aumentar_letra(letra)
                sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
                sheet[letra+str(fila)].font = fuente
                sheet[letra+str(fila)].number_format = formato_numero
                sheet[letra+str(fila)] = det['cheques_postfechados']
                total_cheques_postfechados += det['cheques_postfechados']

                if v_saldo:
                    letra = aumentar_letra(letra)
                    sheet[letra+str(fila)].alignment = Alignment(wrapText=True, horizontal='right', vertical='top')
                    sheet[letra+str(fila)].font = fuente
                    sheet[letra+str(fila)].number_format = formato_numero
                    sheet[letra+str(fila)] = det["saldo"]

            
            if tutor:
                letra = "F"
            else:
                letra = aumentar_letra(letra)
            sheet[letra+str(fila)].alignment = Alignment(wrapText=True, horizontal='left', vertical='top')
            sheet[letra+str(fila)].font = fuente
            sheet[letra+str(fila)] = det['comentario']

            saldo += float(det['saldo'])
            total += float(det['total'])
            fila += 1

        if tutor:
            page_break = Break(id=fila)  # create Break obj
            sheet.row_breaks.append(page_break)  # insert page break

        if not tutor:
            #total
            sheet['F'+str(fila)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
            sheet['F'+str(fila)].font = fuente2
            sheet['F'+str(fila)] = "TOTAL"

            sheet['G'+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
            sheet['G'+str(fila)].font = fuente2
            sheet['G'+str(fila)] = total_dias_mora

            letra = "H"
            if v_total:
                sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
                sheet[letra+str(fila)].font = fuente2
                sheet[letra+str(fila)].number_format = formato_numero
                sheet[letra+str(fila)] = total
                letra = "I"

            sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
            sheet[letra+str(fila)].font = fuente2
            sheet[letra+str(fila)].number_format = formato_numero
            sheet[letra+str(fila)] = total_pagos

            letra = aumentar_letra(letra)
            sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
            sheet[letra+str(fila)].font = fuente2
            sheet[letra+str(fila)].number_format = formato_numero
            sheet[letra+str(fila)] = total_cheques_postfechados

            if v_saldo:
                letra = aumentar_letra(letra)
                sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
                sheet[letra+str(fila)].font = fuente2
                sheet[letra+str(fila)].number_format = formato_numero
                sheet[letra+str(fila)] = saldo
                letra = "I"

            total_general += total
            saldo_general += saldo
            total_general_cheques_postfechados += total_cheques_postfechados
            total_general_dias_mora += total_dias_mora
            total_general_pagos += total_pagos

            fila= fila + 2

    if not tutor:
        #total general
        sheet['F'+str(fila)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['F'+str(fila)].font = fuente2
        sheet['F'+str(fila)] = 'Total General'

        sheet['G'+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
        sheet['G'+str(fila)].font = fuente2
        sheet['G'+str(fila)] = total_general_dias_mora

        letra = "H"
        if v_total:
            sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
            sheet[letra+str(fila)].font = fuente2
            sheet[letra+str(fila)].number_format = formato_numero
            sheet[letra+str(fila)] = total_general
            letra = "I"

        sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
        sheet[letra+str(fila)].font = fuente2
        sheet[letra+str(fila)].number_format = formato_numero
        sheet[letra+str(fila)] = total_general_pagos

        letra = aumentar_letra(letra)
        sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
        sheet[letra+str(fila)].font = fuente2
        sheet[letra+str(fila)].number_format = formato_numero
        sheet[letra+str(fila)] = total_general_cheques_postfechados

        if v_saldo:
            letra = aumentar_letra(letra)
            sheet[letra+str(fila)].alignment = Alignment(wrapText=True,horizontal='right', vertical='top')
            sheet[letra+str(fila)].font = fuente2
            sheet[letra+str(fila)].number_format  = formato_numero
            sheet[letra+str(fila)] = saldo_general


        sheet.merge_cells('A'+str(fila)+':B'+str(fila))
        sheet['A'+str(fila)].alignment = Alignment(wrapText=True,horizontal='center', vertical='top')
        sheet['A'+str(fila)].font = fuente2
        sheet['A'+str(fila)] = 'Elaborado por:'

        sheet.merge_cells('C'+str(fila)+':D'+str(fila))
        sheet['C'+str(fila)].alignment = Alignment(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila)].font = fuente
        sheet['C'+str(fila)] = str(dic['usuario_id'])
