# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from copy import deepcopy
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, date, timedelta
import time
import locale
global root

def crear_wb_informe():
    wb = openpyxl.Workbook()
    return wb


def unicodeText(text):
    try:
        text = unicodedata.unicode(text, 'utf-8')
    except TypeError:
        return text

def crea_hoja_info(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet

def crea_hoja_info_pdf(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet



def border_tabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

    colfin=colfin+1
    filfin=filfin+2

    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    for i in range(fil, filfin-1):
        for j in range(col, colfin):
            sheet.cell(row=i, column=j).border = border_cell


def columnas_filas(sheet, flag, celda, value):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value



def poner_border(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    sheet.cell(row=fil, column=col).border = border_cell

def Informe(sheet, dic,lista_alumnos,cant_alumno,filtro):
    columnas_filas(sheet, 0, 'A', 10.00)
    columnas_filas(sheet, 0, 'B', 5.00)
    columnas_filas(sheet, 0, 'C', 10.00)     
    columnas_filas(sheet, 0, 'D', 7.00)
    columnas_filas(sheet, 0, 'E', 12.00)
    columnas_filas(sheet, 0, 'F', 10.00)
    columnas_filas(sheet, 0, 'G', 10.00)
    columnas_filas(sheet, 0, 'H', 7.00)
    columnas_filas(sheet, 0, 'I', 10.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=6, name='arial')
    fuente3 = Font(bold=True, size=8, name='arial')
    fuente2 = Font(bold=True, size=6, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:I2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente3
    sheet['A2']= 'REPORTE DE CAJA'

    sheet['H1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['H1'].font = fuente2
    sheet['H1']= 'Usuario'
    usuario_id=str(dic['usuario_id'].encode('utf-8'))
    sheet['I1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I1'].font = fuente
    sheet['I1']= str(dic['usuario_id'].encode('utf-8'))

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente2
    sheet['A1']= 'Cia'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente
    sheet['B1']= str(dic['company_id'].encode('utf-8'))

    sheet['A3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A3'].font = fuente2
    sheet['A3']= 'Fecha Emision:'
    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha_corte']

    sheet.merge_cells('B3:C3')
    sheet['B3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B3'].font = fuente
    sheet['B3']= fecha_actual

    poner_border(sheet,1,1,'medium','medium','none','none')
    poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'none','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','none','none')
    poner_border(sheet,1,7,'none','medium','none','none')
    poner_border(sheet,1,8,'none','medium','none','none')
    poner_border(sheet,1,9,'none','medium','medium','none')

    poner_border(sheet,2,1,'medium','none','none','none')
    poner_border(sheet,2,9,'none','none','medium','none')

    poner_border(sheet,3,1,'medium','none','none','medium')
    poner_border(sheet,3,2,'none','none','none','medium')
    poner_border(sheet,3,3,'none','none','none','medium')
    poner_border(sheet,3,4,'none','none','none','medium')
    poner_border(sheet,3,5,'none','none','none','medium')
    poner_border(sheet,3,6,'none','none','none','medium')
    poner_border(sheet,3,7,'none','none','none','medium')
    poner_border(sheet,3,8,'none','none','none','medium')
    poner_border(sheet,3,9,'none','none','medium','medium')
    fecha_ini=dic['fecha_desde']
    fecha_fin=dic['fecha_hasta']
    fecha=str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])
    sheet.merge_cells('D4:F4')
    sheet['D4'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D4'].font = fuente2
    sheet['D4']= str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])

    sheet['A5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A5'].font = fuente2
    sheet['A5']= 'Origen'

    sheet['B5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B5'].font = fuente2
    sheet['B5']= 'Fecha'

    sheet['C5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C5'].font = fuente2
    sheet['C5']= 'FACTURA'

    sheet['D5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['D5'].font = fuente2
    sheet['D5']= 'MONTO'

    sheet['E5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['E5'].font = fuente2
    sheet['E5']= 'Alumno'

    sheet['F5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F5'].font = fuente2
    sheet['F5']= 'Banco'

    sheet['G5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G5'].font = fuente2
    sheet['G5']= 'DOCUMENTO'

    sheet['H5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H5'].font = fuente2
    sheet['H5']= 'Fecha Cheque'

    sheet['I5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I5'].font = fuente2
    sheet['I5']= 'Comentario'

    fila=6
    total_general=0.0
    saldo_general=0.0
    dic={}
    lista_datos=[]
    for recorrer in lista_alumnos:
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente2
        dic={}
        if recorrer['tipo']=='efe':
            sheet['A'+str(fila)]= 'Efectivo'
            dic['tipo']='Efectivo'
        elif recorrer['tipo']=='ch':
            sheet['A'+str(fila)]= 'Cheque'
            dic['tipo']='Cheque'
        elif recorrer['tipo']=='tc':
            sheet['A'+str(fila)]= 'Tarjeta de Credito'
            dic['tipo']='Tarjeta de Credito'
        elif recorrer['tipo']=='dep':
            sheet['A'+str(fila)]= 'Deposito Bancario'
            dic['tipo']='Deposito Bancario'
        elif recorrer['tipo']=='trans':
            sheet['A'+str(fila)]= 'Transferencia Bancaria'
            dic['tipo']='Transferencia Bancaria'
        elif recorrer['tipo']=='nc':
            sheet['A'+str(fila)]= 'Nota de Credito'
            dic['tipo']='Nota de Credito'
        elif recorrer['tipo']=='rti':
            sheet['A'+str(fila)]= 'Retencion iva'
            dic['tipo']='Retencion iva'
        elif recorrer['tipo']=='rtf':
            sheet['A'+str(fila)]= 'Retencion fuente'
            dic['tipo']='Retencion fuente'
        elif recorrer['tipo']=='liq':
            sheet['A'+str(fila)]= 'Liquidacion'
            dic['tipo']='Liquidacion'

        fila=fila+1
        saldo=0.0
        total=0.0
        dic['cantidad']=len(recorrer['detalle'])
        for det in recorrer['detalle']:
            sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['A'+str(fila)].font = fuente
            sheet['A'+str(fila)]= det['numero']

            sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['B'+str(fila)].font = fuente
            sheet['B'+str(fila)]= det['fecha_pago']

            sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['C'+str(fila)].font = fuente
            sheet['C'+str(fila)]= det['factura']

            sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
            sheet['D'+str(fila)].font = fuente
            sheet['D'+str(fila)]= "{:,}".format(float(det['monto'])).replace(',','~').replace('.',',').replace('~','.')

            sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['E'+str(fila)].font = fuente
            if det['cliente']==False:
                sheet['E'+str(fila)]= ''
            else:
                sheet['E'+str(fila)]= det['cliente']

            sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['F'+str(fila)].font = fuente
            if det['banco']==False:
                sheet['F'+str(fila)]= ''
            elif det['banco']==0:
                sheet['F'+str(fila)]= ''
            else:
                sheet['F'+str(fila)]= det['banco']


            sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['G'+str(fila)].font = fuente
            if det['documento']==False:
                sheet['G'+str(fila)]= ''
            else:
                sheet['G'+str(fila)]= det['documento']

            sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['H'+str(fila)].font = fuente
            if det['fecha_ch']==False:
                sheet['H'+str(fila)]= ''
            else:
                sheet['H'+str(fila)]= det['fecha_ch']

            sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['I'+str(fila)].font = fuente
            if det['comentario']==False:
                sheet['I'+str(fila)]= ''
            elif det['comentario']==0:
                sheet['I'+str(fila)]= ''
            else:
                sheet['I'+str(fila)]= det['comentario']

            total=total+float(det['monto'])
            fila=fila+1

        sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila)].font = fuente2
        sheet['C'+str(fila)]= 'TOTAL'

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila)].font = fuente2
        sheet['D'+str(fila)]= "{:,}".format(float(total)).replace(',','~').replace('.',',').replace('~','.')

        dic['total']=total

        total_general = total_general + total

        fila= fila + 1
        lista_datos.append(dic)

    sheet['C'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C'+str(fila+1)].font = fuente2
    sheet['C'+str(fila+1)]= 'TOTAL GENERAL'

    sheet['D'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['D'+str(fila+1)].font = fuente2
    sheet['D'+str(fila+1)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')

    sheet.merge_cells('B'+str(fila+2)+':C'+str(fila+2))
    sheet['B'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B'+str(fila+2)].font = fuente2
    sheet['B'+str(fila+2)]= 'RESUMEN DE VALORES'

    fila=fila+3
    for d in lista_datos:
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente2
        sheet['A'+str(fila)]= d['tipo']

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila)].font = fuente2
        sheet['D'+str(fila)]= "{:,}".format(float(d['total'])).replace(',','~').replace('.',',').replace('~','.')

        sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['H'+str(fila)].font = fuente2
        sheet['H'+str(fila)]= d['cantidad']

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['I'+str(fila)].font = fuente2
        sheet['I'+str(fila)]= 'VECES'

        fila=fila+1

    sheet['D'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['D'+str(fila+1)].font = fuente2
    sheet['D'+str(fila+1)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')

    columnas_filas(sheet, 1, str(fila+5), 10.00)
    sheet.merge_cells('D'+str(fila+5)+':F'+str(fila+5))
    sheet['D'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['D'+str(fila+5)].font = fuente2
    sheet['D'+str(fila+5)]= usuario_id
    poner_border(sheet,fila+5,4,'none','thin','none','none')
    poner_border(sheet,fila+5,5,'none','thin','none','none')
    poner_border(sheet,fila+5,6,'none','thin','none','none')

    columnas_filas(sheet, 1, str(fila+6), 8.00)
    sheet.merge_cells('D'+str(fila+6)+':F'+str(fila+6))
    sheet['D'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['D'+str(fila+6)].font = fuente2
    sheet['D'+str(fila+6)]= fecha

    columnas_filas(sheet, 1, str(fila+7), 8.00)
    sheet.merge_cells('D'+str(fila+7)+':F'+str(fila+7))
    sheet['D'+str(fila+7)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['D'+str(fila+7)].font = fuente2
    sheet['D'+str(fila+7)]= filtro


def Informe_pdf(sheet, dic,lista_alumnos,cant_alumno,filtro):
    columnas_filas(sheet, 0, 'A', 10.00)
    columnas_filas(sheet, 0, 'B', 5.00)
    columnas_filas(sheet, 0, 'C', 10.00)     
    columnas_filas(sheet, 0, 'D', 7.00)
    columnas_filas(sheet, 0, 'E', 1.00)
    columnas_filas(sheet, 0, 'F', 12.00)
    columnas_filas(sheet, 0, 'G', 10.00)
    columnas_filas(sheet, 0, 'H', 10.00)
    columnas_filas(sheet, 0, 'I', 7.00)
    columnas_filas(sheet, 0, 'J', 10.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=6, name='arial')
    fuente3 = Font(bold=True, size=8, name='arial')
    fuente2 = Font(bold=True, size=6, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:I2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente3
    sheet['A2']= 'REPORTE DE CAJA'

    sheet['H1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['H1'].font = fuente2
    sheet['H1']= 'Usuario'
    usuario_id=str(dic['usuario_id'].encode('utf-8'))
    sheet['I1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I1'].font = fuente
    sheet['I1']= str(dic['usuario_id'].encode('utf-8'))

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente2
    sheet['A1']= 'Cia'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente
    sheet['B1']= str(dic['company_id'].encode('utf-8'))

    sheet['A3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A3'].font = fuente2
    sheet['A3']= 'Fecha Emision:'
    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha_corte']

    sheet.merge_cells('B3:C3')
    sheet['B3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B3'].font = fuente
    sheet['B3']= fecha_actual

    poner_border(sheet,1,1,'medium','medium','none','none')
    poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'none','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','none','none')
    poner_border(sheet,1,7,'none','medium','none','none')
    poner_border(sheet,1,8,'none','medium','none','none')
    poner_border(sheet,1,9,'none','medium','none','none')
    poner_border(sheet,1,10,'none','medium','medium','none')

    poner_border(sheet,2,1,'medium','none','none','none')
    poner_border(sheet,2,10,'none','none','medium','none')

    poner_border(sheet,3,1,'medium','none','none','medium')
    poner_border(sheet,3,2,'none','none','none','medium')
    poner_border(sheet,3,3,'none','none','none','medium')
    poner_border(sheet,3,4,'none','none','none','medium')
    poner_border(sheet,3,5,'none','none','none','medium')
    poner_border(sheet,3,6,'none','none','none','medium')
    poner_border(sheet,3,7,'none','none','none','medium')
    poner_border(sheet,3,8,'none','none','none','medium')
    poner_border(sheet,3,9,'none','none','none','medium')
    poner_border(sheet,3,10,'none','none','medium','medium')
    fecha_ini=dic['fecha_desde']
    fecha_fin=dic['fecha_hasta']
    fecha=str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])
    sheet.merge_cells('D4:F4')
    sheet['D4'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D4'].font = fuente2
    sheet['D4']= str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])

    sheet['A5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A5'].font = fuente2
    sheet['A5']= 'Origen'

    sheet['B5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B5'].font = fuente2
    sheet['B5']= 'Fecha'

    sheet['C5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C5'].font = fuente2
    sheet['C5']= 'FACTURA'

    sheet['D5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['D5'].font = fuente2
    sheet['D5']= 'MONTO'

    sheet['F5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F5'].font = fuente2
    sheet['F5']= 'Alumno'

    sheet['G5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G5'].font = fuente2
    sheet['G5']= 'Banco'

    sheet['H5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H5'].font = fuente2
    sheet['H5']= 'DOCUMENTO'

    sheet['I5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I5'].font = fuente2
    sheet['I5']= 'Fecha Cheque'

    sheet['J5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['J5'].font = fuente2
    sheet['J5']= 'Comentario'

    fila=6
    total_general=0.0
    saldo_general=0.0
    dic={}
    lista_datos=[]
    for recorrer in lista_alumnos:
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente2
        dic={}
        if recorrer['tipo']=='efe':
            sheet['A'+str(fila)]= 'Efectivo'
            dic['tipo']='Efectivo'
        elif recorrer['tipo']=='ch':
            sheet['A'+str(fila)]= 'Cheque'
            dic['tipo']='Cheque'
        elif recorrer['tipo']=='tc':
            sheet['A'+str(fila)]= 'Tarjeta de Credito'
            dic['tipo']='Tarjeta de Credito'
        elif recorrer['tipo']=='dep':
            sheet['A'+str(fila)]= 'Deposito Bancario'
            dic['tipo']='Deposito Bancario'
        elif recorrer['tipo']=='trans':
            sheet['A'+str(fila)]= 'Transferencia Bancaria'
            dic['tipo']='Transferencia Bancaria'
        elif recorrer['tipo']=='nc':
            sheet['A'+str(fila)]= 'Nota de Credito'
            dic['tipo']='Nota de Credito'
        elif recorrer['tipo']=='rti':
            sheet['A'+str(fila)]= 'Retencion iva'
            dic['tipo']='Retencion iva'
        elif recorrer['tipo']=='rtf':
            sheet['A'+str(fila)]= 'Retencion fuente'
            dic['tipo']='Retencion fuente'
        elif recorrer['tipo']=='liq':
            sheet['A'+str(fila)]= 'Liquidacion'
            dic['tipo']='Liquidacion'

        fila=fila+1
        saldo=0.0
        total=0.0
        dic['cantidad']=len(recorrer['detalle'])
        for det in recorrer['detalle']:
            columnas_filas(sheet, 1, str(fila), 15.00)
            sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['A'+str(fila)].font = fuente
            sheet['A'+str(fila)]= det['numero']

            sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['B'+str(fila)].font = fuente
            sheet['B'+str(fila)]= det['fecha_pago']

            sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['C'+str(fila)].font = fuente
            sheet['C'+str(fila)]= det['factura']

            sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
            sheet['D'+str(fila)].font = fuente
            monto="{:.2f}".format(float(det['monto']))

            sheet['D'+str(fila)].number_format = '"$"#,##0.00' 
            sheet['D'+str(fila)]= "{:,}".format(float(monto)).replace(',','~').replace('.',',').replace('~','.')
            #sheet['D'+str(fila)]= monto

            sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['F'+str(fila)].font = fuente
            if det['cliente']==False:
                sheet['F'+str(fila)]= ''
            elif det['cliente']==0:
                sheet['F'+str(fila)]= ''
            else:
                sheet['F'+str(fila)]= det['cliente']

            sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['G'+str(fila)].font = fuente
            if det['banco']==False:
                sheet['G'+str(fila)]= ''
            elif det['banco']==0:
                sheet['G'+str(fila)]= ''
            else:
                sheet['G'+str(fila)]= det['banco']


            sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['H'+str(fila)].font = fuente
            if det['documento']==False:
                sheet['H'+str(fila)]= ''
            else:
                sheet['H'+str(fila)]= det['documento']

            sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['I'+str(fila)].font = fuente
            if det['fecha_ch']==False:
                sheet['I'+str(fila)]= ''
            else:
                sheet['I'+str(fila)]= det['fecha_ch']

            sheet['J'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['J'+str(fila)].font = fuente
            if det['comentario']==False:
                sheet['J'+str(fila)]= ''
            elif det['comentario']==0:
                sheet['J'+str(fila)]= ''
            else:
                sheet['J'+str(fila)]= det['comentario']

            total=total+float(det['monto'])
            fila=fila+1

        sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila)].font = fuente2
        sheet['C'+str(fila)]= 'TOTAL'

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila)].font = fuente2
        total_1="{:.2f}".format(float(total))
        sheet['D'+str(fila)].number_format = '"$"#,##0.00' 
        sheet['D'+str(fila)]= "{:,}".format(float(total_1)).replace(',','~').replace('.',',').replace('~','.')

        dic['total']=total

        total_general = total_general + total

        fila= fila + 1
        lista_datos.append(dic)

    sheet['C'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C'+str(fila+1)].font = fuente2
    sheet['C'+str(fila+1)]= 'TOTAL GENERAL'

    sheet['D'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['D'+str(fila+1)].font = fuente2
    total_2="{:.2f}".format(float(total_general))
    sheet['D'+str(fila+1)].number_format = '"$"#,##0.00' 
    sheet['D'+str(fila+1)]= "{:,}".format(float(total_2)).replace(',','~').replace('.',',').replace('~','.')

    sheet.merge_cells('B'+str(fila+2)+':C'+str(fila+2))
    sheet['B'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B'+str(fila+2)].font = fuente2
    sheet['B'+str(fila+2)]= 'RESUMEN DE VALORES'

    fila=fila+3

    for d in lista_datos:
        columnas_filas(sheet, 1, str(fila), 10.00)
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente2
        sheet['A'+str(fila)]= d['tipo']

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila)].font = fuente2
        total_3="{:.2f}".format(float(d['total']))
        sheet['D'+str(fila)].number_format = '"$"#,##0.00' 
        sheet['D'+str(fila)]= "{:,}".format(float(total_3)).replace(',','~').replace('.',',').replace('~','.')

        sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['H'+str(fila)].font = fuente2
        sheet['H'+str(fila)]= d['cantidad']

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['I'+str(fila)].font = fuente2
        sheet['I'+str(fila)]= 'VECES'

        fila=fila+1

    sheet['D'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['D'+str(fila+1)].font = fuente2
    total_4="{:.2f}".format(float(total_general))
    sheet['D'+str(fila+1)].number_format = '"$"#,##0.00' 
    sheet['D'+str(fila+1)]= "{:,}".format(float(total_4)).replace(',','~').replace('.',',').replace('~','.')

    columnas_filas(sheet, 1, str(fila+5), 10.00)
    sheet.merge_cells('D'+str(fila+5)+':F'+str(fila+5))
    sheet['D'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['D'+str(fila+5)].font = fuente2
    sheet['D'+str(fila+5)]= usuario_id
    poner_border(sheet,fila+5,4,'none','thin','none','none')
    poner_border(sheet,fila+5,5,'none','thin','none','none')
    poner_border(sheet,fila+5,6,'none','thin','none','none')

    columnas_filas(sheet, 1, str(fila+6), 8.00)
    sheet.merge_cells('D'+str(fila+6)+':F'+str(fila+6))
    sheet['D'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['D'+str(fila+6)].font = fuente2
    sheet['D'+str(fila+6)]= fecha

    columnas_filas(sheet, 1, str(fila+7), 20.00)
    sheet.merge_cells('D'+str(fila+7)+':F'+str(fila+7))
    sheet['D'+str(fila+7)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D'+str(fila+7)].font = fuente2
    sheet['D'+str(fila+7)]= filtro