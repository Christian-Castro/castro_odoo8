# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from copy import deepcopy
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, date, timedelta
import time
import locale
global root

def crear_wb_informe():
    wb = openpyxl.Workbook()
    return wb


def unicodeText(text):
    try:
        text = unicodedata.unicode(text, 'utf-8')
    except TypeError:
        return text

def crea_hoja_info(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet

def crea_hoja_info_pdf(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet



def border_tabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

    colfin=colfin+1
    filfin=filfin+2

    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    for i in range(fil, filfin-1):
        for j in range(col, colfin):
            sheet.cell(row=i, column=j).border = border_cell


def columnas_filas(sheet, flag, celda, value):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value



def poner_border(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    sheet.cell(row=fil, column=col).border = border_cell

def Informe(sheet, dic,lista_alumnos,cant_alumno):
    columnas_filas(sheet, 0, 'A', 10.00)
    columnas_filas(sheet, 0, 'B', 15.00)
    columnas_filas(sheet, 0, 'C', 15.00)     
    columnas_filas(sheet, 0, 'D', 10.00)
    columnas_filas(sheet, 0, 'E', 12.00)
    columnas_filas(sheet, 0, 'F', 18.00)
    columnas_filas(sheet, 0, 'G', 1.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=7, name='arial')
    fuente3 = Font(bold=True, size=10, name='arial')
    fuente2 = Font(bold=True, size=7, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:F2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente3
    sheet['A2']= 'ESTADO DE CUENTA'

    sheet['E1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['E1'].font = fuente2
    sheet['E1']= 'Usuario'

    sheet['F1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F1'].font = fuente
    sheet['F1']= str(dic['usuario_id'].encode('utf-8'))

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente2
    sheet['A1']= 'Cia'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente
    sheet['B1']= str(dic['company_id'].encode('utf-8'))

    sheet['A3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A3'].font = fuente2
    sheet['A3']= 'Fecha Emision:'
    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha_corte']

    sheet.merge_cells('B3:C3')
    sheet['B3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B3'].font = fuente
    sheet['B3']= fecha_actual

    poner_border(sheet,1,1,'medium','medium','none','none')
    poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'none','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','medium','none')

    poner_border(sheet,2,1,'medium','none','none','none')
    poner_border(sheet,2,6,'none','none','medium','none')

    poner_border(sheet,3,1,'medium','none','none','medium')
    poner_border(sheet,3,2,'none','none','none','medium')
    poner_border(sheet,3,3,'none','none','none','medium')
    poner_border(sheet,3,4,'none','none','none','medium')
    poner_border(sheet,3,5,'none','none','none','medium')
    poner_border(sheet,3,6,'none','none','medium','medium')

    fila=2
    total_general=0.0
    saldo_general=0.0
    for recorrer in lista_alumnos:
        sheet['A'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+2)].font = fuente2
        sheet['A'+str(fila+2)]= 'Alumno:'

        sheet.merge_cells('B'+str(fila+2)+':C'+str(fila+2))
        sheet['B'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+2)].font = fuente
        if recorrer['alumno']!=False:
            sheet['B'+str(fila+2)]= str(recorrer['alumno'])

        sheet['A'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+3)].font = fuente2
        sheet['A'+str(fila+3)]= 'Direccion'

        sheet.merge_cells('B'+str(fila+3)+':C'+str(fila+3))
        sheet['B'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+3)].font = fuente
        sheet['B'+str(fila+3)]= str(recorrer['direccion'].encode('utf-8'))

        sheet['A'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+4)].font = fuente2
        sheet['A'+str(fila+4)]= 'Telefono'

        sheet.merge_cells('B'+str(fila+4)+':C'+str(fila+4))
        sheet['B'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+4)].font = fuente
        sheet['B'+str(fila+4)]= (recorrer['telefono'])

        sheet['D'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+2)].font = fuente2
        sheet['D'+str(fila+2)]= 'Representante'

        sheet.merge_cells('E'+str(fila+2)+':F'+str(fila+2))
        sheet['E'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['E'+str(fila+2)].font = fuente
        if recorrer['representante']!=False:
            sheet['E'+str(fila+2)]= str(recorrer['representante'].encode('utf-8'))

        sheet['D'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+3)].font = fuente2
        sheet['D'+str(fila+3)]= 'Cedula o RUC'

        sheet.merge_cells('E'+str(fila+3)+':F'+str(fila+3))
        sheet['E'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['E'+str(fila+3)].font = fuente
        sheet['E'+str(fila+3)]= (recorrer['cedula'])

        sheet.merge_cells('D'+str(fila+5)+':E'+str(fila+5))
        sheet['D'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+5)].font = fuente
        sheet['D'+str(fila+5)]= 'Jornada: '+(recorrer['jornada'])

        sheet['A'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+6)].font = fuente
        sheet['A'+str(fila+6)]= 'STATUS A'

        sheet.merge_cells('D'+str(fila+6)+':E'+str(fila+6))
        sheet['D'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+6)].font = fuente
        sheet['D'+str(fila+6)]= 'Curso: '+(recorrer['curso'])

        sheet['F'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila+5)].font = fuente
        sheet['F'+str(fila+5)]= 'Seccion: '+(recorrer['seccion'])

        sheet['F'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila+6)].font = fuente
        sheet['F'+str(fila+6)]= 'Paralelo: '+(recorrer['paralelo'])

        
        poner_border(sheet,fila+7,6,'none','medium','medium','none')
        poner_border(sheet,fila+7,1,'medium','medium','none','none')
        poner_border(sheet,fila+7,2,'none','medium','none','none')
        poner_border(sheet,fila+7,3,'none','medium','none','none')
        poner_border(sheet,fila+7,4,'none','medium','none','none')
        poner_border(sheet,fila+7,5,'none','medium','none','none')

        sheet.merge_cells('A'+str(fila+7)+':B'+str(fila+7))
        sheet['A'+str(fila+7)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['A'+str(fila+7)].font = fuente2
        sheet['A'+str(fila+7)]= 'Documento'

        sheet['A'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['A'+str(fila+8)].font = fuente2
        sheet['A'+str(fila+8)]= 'Tipo'

        sheet['B'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['B'+str(fila+8)].font = fuente2
        sheet['B'+str(fila+8)]= 'Numero'

        sheet['C'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['C'+str(fila+8)].font = fuente2
        sheet['C'+str(fila+8)]= 'Emision'

        sheet['D'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['D'+str(fila+8)].font = fuente2
        sheet['D'+str(fila+8)]= 'Cargos'

        sheet['E'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['E'+str(fila+8)].font = fuente2
        sheet['E'+str(fila+8)]= 'Abonos'

        sheet['F'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['F'+str(fila+8)].font = fuente2
        sheet['F'+str(fila+8)]= 'Comentario'

        poner_border(sheet,fila+8,6,'none','none','medium','medium')
        poner_border(sheet,fila+8,1,'medium','none','none','medium')
        poner_border(sheet,fila+8,2,'none','none','none','medium')
        poner_border(sheet,fila+8,3,'none','none','none','medium')
        poner_border(sheet,fila+8,4,'none','none','none','medium')
        poner_border(sheet,fila+8,5,'none','none','none','medium')

        sheet.merge_cells('A'+str(fila+9)+':B'+str(fila+9))
        sheet['A'+str(fila+9)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+9)].font = fuente2
        sheet['A'+str(fila+9)]= 'Saldo al '+str(dic['fecha_desde'])

        sheet['D'+str(fila+9)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila+9)].font = fuente2
        sheet['D'+str(fila+9)]= "{:,}".format(float(recorrer['cargo'])).replace(',','~').replace('.',',').replace('~','.')

        sheet['E'+str(fila+9)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['E'+str(fila+9)].font = fuente2
        sheet['E'+str(fila+9)]= 0.00

        fila=fila+10
        saldo=0.0
        total=0.0
        for det in recorrer['detalle']:
            sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['A'+str(fila)].font = fuente
            sheet['A'+str(fila)]= det['tipo']

            sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['B'+str(fila)].font = fuente
            sheet['B'+str(fila)]= det['numero']

            sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
            sheet['C'+str(fila)].font = fuente
            sheet['C'+str(fila)]= det['emision']

            sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
            sheet['D'+str(fila)].font = fuente
            sheet['D'+str(fila)]= "{:,}".format(float(det['cargos'])).replace(',','~').replace('.',',').replace('~','.')

            sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
            sheet['E'+str(fila)].font = fuente
            sheet['E'+str(fila)]= "{:,}".format(float(det['total'])).replace(',','~').replace('.',',').replace('~','.')

            sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['F'+str(fila)].font = fuente
            sheet['F'+str(fila)]= det['comentario']

            saldo=saldo+float(det['cargos'])
            total=total+float(det['total'])
            fila=fila+1

        sheet.merge_cells('A'+str(fila)+':B'+str(fila))
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente2
        sheet['A'+str(fila)]= 'Saldo al '+str(dic['fecha_hasta'])

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila)].font = fuente2
        sheet['D'+str(fila)]= "{:,}".format(float(saldo)).replace(',','~').replace('.',',').replace('~','.')

        sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['E'+str(fila)].font = fuente2
        sheet['E'+str(fila)]= "{:,}".format(float(total)).replace(',','~').replace('.',',').replace('~','.')

        resta=saldo-total

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['F'+str(fila)].font = fuente2
        sheet['F'+str(fila)]= "{:,}".format(float(resta)).replace(',','~').replace('.',',').replace('~','.')

        poner_border(sheet,fila,4,'none','medium','none','none')
        poner_border(sheet,fila,5,'none','medium','none','none')

        total_general = total_general + total
        saldo_general = saldo_general + saldo

        fila= fila + 2

    sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['C'+str(fila)].font = fuente2
    sheet['C'+str(fila)]= 'Total General'

    sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['D'+str(fila)].font = fuente2
    sheet['D'+str(fila)]= "{:,}".format(float(saldo_general)).replace(',','~').replace('.',',').replace('~','.')

    sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['E'+str(fila)].font = fuente2
    sheet['E'+str(fila)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')


def Informe_pdf(sheet, dic,lista_alumnos,cant_alumno):
    columnas_filas(sheet, 0, 'A', 8.00)
    columnas_filas(sheet, 0, 'B', 15.00)
    columnas_filas(sheet, 0, 'C', 15.00)     
    columnas_filas(sheet, 0, 'D', 10.00)
    columnas_filas(sheet, 0, 'E', 12.00)
    columnas_filas(sheet, 0, 'F', 18.00)
    columnas_filas(sheet, 0, 'G', 1.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=7, name='arial')
    fuente3 = Font(bold=True, size=10, name='arial')
    fuente2 = Font(bold=True, size=7, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:F2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente3
    sheet['A2']= 'ESTADO DE CUENTA'

    sheet['E1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['E1'].font = fuente2
    sheet['E1']= 'Usuario'

    sheet['F1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F1'].font = fuente
    sheet['F1']= str(dic['usuario_id'].encode('utf-8'))

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente2
    sheet['A1']= 'Cia'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente
    sheet['B1']= str(dic['company_id'].encode('utf-8'))

    sheet['A3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A3'].font = fuente2
    sheet['A3']= 'Fecha Emision:'
    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha_corte']

    sheet.merge_cells('B3:C3')
    sheet['B3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B3'].font = fuente
    sheet['B3']= fecha_actual

    poner_border(sheet,1,1,'medium','medium','none','none')
    poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'none','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','medium','none')

    poner_border(sheet,2,1,'medium','none','none','none')
    poner_border(sheet,2,6,'none','none','medium','none')

    poner_border(sheet,3,1,'medium','none','none','medium')
    poner_border(sheet,3,2,'none','none','none','medium')
    poner_border(sheet,3,3,'none','none','none','medium')
    poner_border(sheet,3,4,'none','none','none','medium')
    poner_border(sheet,3,5,'none','none','none','medium')
    poner_border(sheet,3,6,'none','none','medium','medium')

    fila=2
    total_general=0.0
    saldo_general=0.0
    for recorrer in lista_alumnos:
        columnas_filas(sheet, 1, str(fila+2), 10.00)
        columnas_filas(sheet, 1, str(fila+3), 15.00)
        columnas_filas(sheet, 1, str(fila+4), 10.00)
        columnas_filas(sheet, 1, str(fila+5), 10.00)
        columnas_filas(sheet, 1, str(fila+6), 10.00)
        columnas_filas(sheet, 1, str(fila+7), 10.00)
        columnas_filas(sheet, 1, str(fila+8), 10.00)
        columnas_filas(sheet, 1, str(fila+9), 10.00)
        sheet['A'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+2)].font = fuente2
        sheet['A'+str(fila+2)]= 'Alumno:'

        sheet.merge_cells('B'+str(fila+2)+':C'+str(fila+2))
        sheet['B'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
        sheet['B'+str(fila+2)].font = fuente
        if recorrer['alumno']!=False:
            sheet['B'+str(fila+2)]= str(recorrer['alumno'])

        sheet['A'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+3)].font = fuente2
        sheet['A'+str(fila+3)]= 'Direccion'

        sheet.merge_cells('B'+str(fila+3)+':C'+str(fila+3))
        sheet['B'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='justify')
        sheet['B'+str(fila+3)].font = fuente
        sheet['B'+str(fila+3)]= str(recorrer['direccion'].encode('utf-8'))

        sheet['A'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+4)].font = fuente2
        sheet['A'+str(fila+4)]= 'Telefono'

        sheet.merge_cells('B'+str(fila+4)+':C'+str(fila+4))
        sheet['B'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
        sheet['B'+str(fila+4)].font = fuente
        sheet['B'+str(fila+4)]= (recorrer['telefono'])

        sheet['D'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+2)].font = fuente2
        sheet['D'+str(fila+2)]= 'Representante'

        sheet.merge_cells('E'+str(fila+2)+':F'+str(fila+2))
        sheet['E'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
        sheet['E'+str(fila+2)].font = fuente
        if recorrer['representante']!=False:
            sheet['E'+str(fila+2)]= str(recorrer['representante'].encode('utf-8'))

        sheet['D'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+3)].font = fuente2
        sheet['D'+str(fila+3)]= 'Cedula o RUC'

        sheet.merge_cells('E'+str(fila+3)+':F'+str(fila+3))
        sheet['E'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
        sheet['E'+str(fila+3)].font = fuente
        sheet['E'+str(fila+3)]= (recorrer['cedula'])

        sheet.merge_cells('D'+str(fila+5)+':E'+str(fila+5))
        sheet['D'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+5)].font = fuente
        sheet['D'+str(fila+5)]= 'Jornada: '+(recorrer['jornada'])

        sheet['A'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+6)].font = fuente
        sheet['A'+str(fila+6)]= 'STATUS A'

        sheet.merge_cells('D'+str(fila+6)+':E'+str(fila+6))
        sheet['D'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila+6)].font = fuente
        sheet['D'+str(fila+6)]= 'Curso: '+(recorrer['curso'])

        sheet['F'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila+5)].font = fuente
        sheet['F'+str(fila+5)]= 'Seccion: '+(recorrer['seccion'])

        sheet['F'+str(fila+6)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila+6)].font = fuente
        sheet['F'+str(fila+6)]= 'Paralelo: '+(recorrer['paralelo'])

        
        poner_border(sheet,fila+7,6,'none','medium','medium','none')
        poner_border(sheet,fila+7,1,'medium','medium','none','none')
        poner_border(sheet,fila+7,2,'none','medium','none','none')
        poner_border(sheet,fila+7,3,'none','medium','none','none')
        poner_border(sheet,fila+7,4,'none','medium','none','none')
        poner_border(sheet,fila+7,5,'none','medium','none','none')

        sheet.merge_cells('A'+str(fila+7)+':B'+str(fila+7))
        sheet['A'+str(fila+7)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['A'+str(fila+7)].font = fuente2
        sheet['A'+str(fila+7)]= 'Documento'

        sheet['A'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['A'+str(fila+8)].font = fuente2
        sheet['A'+str(fila+8)]= 'Tipo'

        sheet['B'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['B'+str(fila+8)].font = fuente2
        sheet['B'+str(fila+8)]= 'Numero'

        sheet['C'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['C'+str(fila+8)].font = fuente2
        sheet['C'+str(fila+8)]= 'Emision'

        sheet['D'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['D'+str(fila+8)].font = fuente2
        sheet['D'+str(fila+8)]= 'Cargos'

        sheet['E'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['E'+str(fila+8)].font = fuente2
        sheet['E'+str(fila+8)]= 'Abonos'

        sheet['F'+str(fila+8)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['F'+str(fila+8)].font = fuente2
        sheet['F'+str(fila+8)]= 'Comentario'

        poner_border(sheet,fila+8,6,'none','none','medium','medium')
        poner_border(sheet,fila+8,1,'medium','none','none','medium')
        poner_border(sheet,fila+8,2,'none','none','none','medium')
        poner_border(sheet,fila+8,3,'none','none','none','medium')
        poner_border(sheet,fila+8,4,'none','none','none','medium')
        poner_border(sheet,fila+8,5,'none','none','none','medium')

        sheet.merge_cells('A'+str(fila+9)+':B'+str(fila+9))
        sheet['A'+str(fila+9)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+9)].font = fuente2
        sheet['A'+str(fila+9)]= 'Saldo al '+str(dic['fecha_desde'])

        sheet['D'+str(fila+9)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila+9)].font = fuente2
        sheet['D'+str(fila+9)]= "{:,}".format(float(recorrer['cargo'])).replace(',','~').replace('.',',').replace('~','.')

        sheet['E'+str(fila+9)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['E'+str(fila+9)].font = fuente2
        sheet['E'+str(fila+9)]= 0.00

        fila=fila+10
        saldo=0.0
        total=0.0
        for det in recorrer['detalle']:
            columnas_filas(sheet, 1, str(fila), 10.00)
            sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['A'+str(fila)].font = fuente
            sheet['A'+str(fila)]= det['tipo']

            sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['B'+str(fila)].font = fuente
            sheet['B'+str(fila)]= det['numero']

            sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
            sheet['C'+str(fila)].font = fuente
            sheet['C'+str(fila)]= det['emision']

            sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
            sheet['D'+str(fila)].font = fuente
            sheet['D'+str(fila)]= "{:,}".format(float(det['cargos'])).replace(',','~').replace('.',',').replace('~','.')

            sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
            sheet['E'+str(fila)].font = fuente
            sheet['E'+str(fila)]= "{:,}".format(float(det['total'])).replace(',','~').replace('.',',').replace('~','.')

            sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['F'+str(fila)].font = fuente
            sheet['F'+str(fila)]= det['comentario']

            saldo=saldo+float(det['cargos'])
            total=total+float(det['total'])
            fila=fila+1

        sheet.merge_cells('A'+str(fila)+':B'+str(fila))
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente2
        sheet['A'+str(fila)]= 'Saldo al '+str(dic['fecha_hasta'])

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['D'+str(fila)].font = fuente2
        sheet['D'+str(fila)]= "{:,}".format(float(saldo)).replace(',','~').replace('.',',').replace('~','.')

        sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['E'+str(fila)].font = fuente2 
        sheet['E'+str(fila)]= "{:,}".format(float(total)).replace(',','~').replace('.',',').replace('~','.')

        resta=saldo-total

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['F'+str(fila)].font = fuente2
        sheet['F'+str(fila)]= "{:,}".format(float(resta)).replace(',','~').replace('.',',').replace('~','.')

        poner_border(sheet,fila,4,'none','medium','none','none')
        poner_border(sheet,fila,5,'none','medium','none','none')

        total_general = total_general + total
        saldo_general = saldo_general + saldo

        fila= fila + 2

    sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['C'+str(fila)].font = fuente2
    sheet['C'+str(fila)]= 'Total General'

    sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['D'+str(fila)].font = fuente2
    sheet['D'+str(fila)]= "{:,}".format(float(saldo_general)).replace(',','~').replace('.',',').replace('~','.')

    sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['E'+str(fila)].font = fuente2
    sheet['E'+str(fila)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')