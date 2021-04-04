# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from copy import deepcopy
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, date, timedelta
import time
import locale
global root

def crear_wb_informe():
    wb = openpyxl.Workbook()
    return wb


def unicodeText(text):
    try:
        text = unicodedata.unicode(text, 'utf-8')
    except TypeError:
        return text

def crea_hoja_info(wb, title, flag):
    if(flag == 0):
        sheet = wb.active
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
    if(flag == 1):
        sheet = wb.create_sheet()
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
    sheet.title = title
    return sheet

def crea_hoja_info_pdf(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet



def border_tabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

    colfin=colfin+1
    filfin=filfin+2

    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    for i in range(fil, filfin-1):
        for j in range(col, colfin):
            sheet.cell(row=i, column=j).border = border_cell


def columnas_filas(sheet, flag, celda, value):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value



def poner_border(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    sheet.cell(row=fil, column=col).border = border_cell

def Informe(sheet, dic,lista_alumnos,cant_alumno):
    columnas_filas(sheet, 0, 'A', 20.00)
    columnas_filas(sheet, 0, 'B', 18.00)
    columnas_filas(sheet, 0, 'C', 40.00)     
    columnas_filas(sheet, 0, 'D', 40.00)
    columnas_filas(sheet, 0, 'F', 40.00)
    columnas_filas(sheet, 0, 'E', 50.00)
    columnas_filas(sheet, 0, 'G', 20.00)
    columnas_filas(sheet, 0, 'H', 20.00)
    columnas_filas(sheet, 0, 'I', 15.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=11, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:I2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente.copy(bold=True,size=15, name= "arial")
    sheet['A2']= 'DIRECTORIO DE ALUMNOS'

    sheet['G1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['G1'].font = fuente.copy(bold=True,size=12, name= "arial")
    sheet['G1']= 'Usuario'

    sheet['H1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H1'].font = fuente
    sheet['H1']= dic['usuario_id']

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente.copy(bold=True,size=12, name= "arial")
    sheet['A1']= 'Cia'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente.copy(bold=False,size=12, name= "arial")
    sheet['B1']= dic['company_id']

    sheet['A3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A3'].font = fuente.copy(bold=True,size=12, name= "arial")
    sheet['A3']= 'Fecha Emision:'
    

    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha']

    sheet.merge_cells('B3:C3')
    sheet['B3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B3'].font = fuente.copy(bold=True,size=12, name= "arial")
    sheet['B3']= fecha_actual

    fila=6
    for recorrer in lista_alumnos:
        sheet['A'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+1)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['A'+str(fila+1)]= 'Seccion:'

        sheet.merge_cells('B'+str(fila+1)+':C'+str(fila+1))
        sheet['B'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+1)].font = fuente.copy(bold=False,size=12, name= "arial")
        if recorrer['seccion_id']==False:
            sheet['B'+str(fila+1)]= ''    
        else:
            sheet['B'+str(fila+1)]= recorrer['seccion_id']

        sheet['A'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+2)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['A'+str(fila+2)]= 'Curso:'

        sheet.merge_cells('B'+str(fila+2)+':C'+str(fila+2))
        sheet['B'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+2)].font = fuente.copy(bold=False,size=12, name= "arial")
        if recorrer['curso_id']==False:
            sheet['B'+str(fila+2)]= ''    
        else:
            sheet['B'+str(fila+2)]= recorrer['curso_id']


        sheet['A'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+3)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['A'+str(fila+3)]= 'Codigo:'

        sheet.merge_cells('B'+str(fila+3)+':C'+str(fila+3))
        sheet['B'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+3)].font = fuente.copy(bold=False,size=12, name= "arial")
        if recorrer['codigo_curso']==False:
            sheet['B'+str(fila+3)]= ''    
        else:
            sheet['B'+str(fila+3)]= recorrer['codigo_curso']


        sheet['A'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+4)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['A'+str(fila+4)]= 'Paralelo:'

        sheet.merge_cells('B'+str(fila+4)+':C'+str(fila+4))
        sheet['B'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+4)].font = fuente.copy(bold=False,size=12, name= "arial")
        if recorrer['paralelo_id']==False:
            sheet['B'+str(fila+4)]= ''    
        else:
            sheet['B'+str(fila+4)]= recorrer['paralelo_id']


        sheet['A'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+5)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['A'+str(fila+5)]= 'Total Alumnos:'

        sheet['B'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila+5)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['B'+str(fila+5)]= len(recorrer['lista_detalle'])

        fila=fila+7

        #poner_border(sheet,13,7,'medium','none','none','none')
        border_tabla(sheet,1,9,fila,fila,'none','none','none','medium')

        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['A'+str(fila)]= 'Orden'

        sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['B'+str(fila)]= 'Codigo'

        sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['C'+str(fila)]= 'Alumno'

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['D'+str(fila)]= 'Representante'

        sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['E'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['E'+str(fila)]= 'Correo'

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['F'+str(fila)]= 'Direccion'

        sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['G'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['G'+str(fila)]= 'Telefono'

        sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['H'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['H'+str(fila)]= 'Cedula - RUC'

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['I'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
        sheet['I'+str(fila)]= 'Status'

        fila=fila+1
        orden=1
        for detalle in recorrer['lista_detalle']:       
            sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['A'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            sheet['A'+str(fila)]= orden
            orden=orden+1

            sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['B'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            sheet['B'+str(fila)]= detalle['codigo']

            sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['C'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            sheet['C'+str(fila)]= detalle['alumno']

            sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['D'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            sheet['D'+str(fila)]= detalle['representante']

            sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['E'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            if detalle['correo']==False:
                sheet['E'+str(fila)]= ''    
            else:
                sheet['E'+str(fila)]= detalle['correo']

            sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['F'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            if detalle['direccion']==False:
                sheet['F'+str(fila)]= ''    
            else:
                sheet['F'+str(fila)]= detalle['direccion']

            sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
            sheet['G'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            if detalle['telefono']==False:
                sheet['G'+str(fila)]= ''    
            else:
                sheet['G'+str(fila)]= detalle['telefono']

            sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['H'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            sheet['H'+str(fila)]= detalle['cedula']

            sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['I'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
            sheet['I'+str(fila)]= detalle['state']

            fila=fila+1
        border_tabla(sheet,1,9,fila,fila,'none','none','none','medium')
        fila=fila+2




    fila=fila+2
    sheet.merge_cells('A'+str(fila)+':B'+str(fila))
    sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A'+str(fila)].font = fuente.copy(bold=True,size=12, name= "arial")
    sheet['A'+str(fila)]= 'TOTAL ALUMNOS REPORTE'


    sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C'+str(fila)].font = fuente.copy(bold=False,size=12, name= "arial")
    sheet['C'+str(fila)]= cant_alumno

def Informe_pdf(sheet, dic,lista_alumnos,cant_alumno):
    columnas_filas(sheet, 0, 'A', 4.00)
    columnas_filas(sheet, 0, 'B', 5.00)
    columnas_filas(sheet, 0, 'C', 17.00)     
    columnas_filas(sheet, 0, 'D', 17.00)
    columnas_filas(sheet, 0, 'F', 10.00)
    columnas_filas(sheet, 0, 'E', 10.00)
    columnas_filas(sheet, 0, 'G', 8.00)
    columnas_filas(sheet, 0, 'H', 7.00)
    columnas_filas(sheet, 0, 'I', 4.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=7, name='arial')
    fuente3 = Font(bold=True, size=7, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:I2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente.copy(bold=True,size=10, name= "arial")
    sheet['A2']= 'DIRECTORIO DE ALUMNOS'

    sheet['G1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['G1'].font = fuente3
    sheet['G1']= 'Usuario:'

    sheet.merge_cells('H1:I1')
    sheet['H1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H1'].font = fuente
    sheet['H1']= dic['usuario_id']

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente3
    sheet['A1']= 'Cia:'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente
    sheet['B1']= dic['company_id']

    sheet.merge_cells('A3:B3')
    sheet['A3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A3'].font = fuente3
    sheet['A3']= 'Fecha Emisión:'
    

    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha']

    sheet.merge_cells('C3:D3')
    sheet['C3'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C3'].font = fuente3
    sheet['C3']= fecha_actual

    fila=6
    for recorrer in lista_alumnos:
        sheet.merge_cells('A'+str(fila+1)+':B'+str(fila+1))
        sheet['A'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+1)].font = fuente3
        sheet['A'+str(fila+1)]= 'Sección:'

        sheet.merge_cells('C'+str(fila+1)+':D'+str(fila+1))
        sheet['C'+str(fila+1)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila+1)].font = fuente
        if recorrer['seccion_id']==False:
            sheet['C'+str(fila+1)]= ''    
        else:
            sheet['C'+str(fila+1)]= recorrer['seccion_id']

        sheet.merge_cells('A'+str(fila+2)+':B'+str(fila+2))
        sheet['A'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+2)].font = fuente3
        sheet['A'+str(fila+2)]= 'Curso:'

        sheet.merge_cells('C'+str(fila+2)+':D'+str(fila+2))
        sheet['C'+str(fila+2)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila+2)].font = fuente
        if recorrer['curso_id']==False:
            sheet['C'+str(fila+2)]= ''    
        else:
            sheet['C'+str(fila+2)]= recorrer['curso_id']

        sheet.merge_cells('A'+str(fila+3)+':B'+str(fila+3))
        sheet['A'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+3)].font = fuente3
        sheet['A'+str(fila+3)]= 'Código:'

        sheet.merge_cells('C'+str(fila+3)+':D'+str(fila+3))
        sheet['C'+str(fila+3)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila+3)].font = fuente
        if recorrer['codigo_curso']==False:
            sheet['C'+str(fila+3)]= ''    
        else:
            sheet['C'+str(fila+3)]= recorrer['codigo_curso']

        sheet.merge_cells('A'+str(fila+4)+':B'+str(fila+4))
        sheet['A'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+4)].font = fuente3
        sheet['A'+str(fila+4)]= 'Paralelo:'

        sheet.merge_cells('C'+str(fila+4)+':D'+str(fila+4))
        sheet['C'+str(fila+4)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila+4)].font = fuente
        if recorrer['paralelo_id']==False:
            sheet['C'+str(fila+4)]= ''    
        else:
            sheet['C'+str(fila+4)]= recorrer['paralelo_id']

        sheet.merge_cells('A'+str(fila+5)+':B'+str(fila+5))
        sheet['A'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila+5)].font = fuente3
        sheet['A'+str(fila+5)]= 'Total Alumnos:'

        sheet['C'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila+5)].font = fuente3
        sheet['C'+str(fila+5)]= len(recorrer['lista_detalle'])

        fila=fila+7

        #poner_border(sheet,13,7,'medium','none','none','none')
        border_tabla(sheet,1,9,fila,fila,'none','none','none','medium')

        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente3
        sheet['A'+str(fila)]= 'Orden'

        sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila)].font = fuente3
        sheet['B'+str(fila)]= 'Código'

        sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['C'+str(fila)].font = fuente3
        sheet['C'+str(fila)]= 'Alumno'

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila)].font = fuente3
        sheet['D'+str(fila)]= 'Representante'

        sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['E'+str(fila)].font = fuente3
        sheet['E'+str(fila)]= 'Correo'

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila)].font = fuente3
        sheet['F'+str(fila)]= 'Dirección'

        sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['G'+str(fila)].font = fuente3
        sheet['G'+str(fila)]= 'Telefono'

        sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['H'+str(fila)].font = fuente3
        sheet['H'+str(fila)]= 'Cédula - RUC'

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['I'+str(fila)].font = fuente3
        sheet['I'+str(fila)]= 'Status'

        fila=fila+1
        orden=1
        for detalle in recorrer['lista_detalle']:       
            sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['A'+str(fila)].font = fuente
            sheet['A'+str(fila)]= orden
            orden=orden+1

            sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['B'+str(fila)].font = fuente
            sheet['B'+str(fila)]= detalle['codigo']

            sheet['C'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['C'+str(fila)].font = fuente
            sheet['C'+str(fila)]= detalle['alumno']

            sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['D'+str(fila)].font = fuente
            sheet['D'+str(fila)]= detalle['representante']

            sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['E'+str(fila)].font = fuente
            if detalle['correo']==False:
                sheet['E'+str(fila)]= ''    
            else:
                sheet['E'+str(fila)]= detalle['correo']

            sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='justify')
            sheet['F'+str(fila)].font = fuente
            if detalle['direccion']==False:
                sheet['F'+str(fila)]= ''    
            else:
                sheet['F'+str(fila)]= detalle['direccion']

            sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['G'+str(fila)].font = fuente
            if detalle['telefono']==False:
                sheet['G'+str(fila)]= ''    
            else:
                sheet['G'+str(fila)]= detalle['telefono']

            sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['H'+str(fila)].font = fuente
            sheet['H'+str(fila)]= detalle['cedula']

            sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
            sheet['I'+str(fila)].font = fuente
            sheet['I'+str(fila)]= detalle['state']

            fila=fila+1
        border_tabla(sheet,1,9,fila,fila,'none','none','none','medium')
        fila=fila+2


    fila=fila+2
    sheet.merge_cells('A'+str(fila)+':C'+str(fila))
    sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A'+str(fila)].font = fuente3
    sheet['A'+str(fila)]= 'TOTAL ALUMNOS REPORTE'


    sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['D'+str(fila)].font = fuente
    sheet['D'+str(fila)]= cant_alumno