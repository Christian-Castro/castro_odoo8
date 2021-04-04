# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from copy import deepcopy
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, date, timedelta
import time
import locale
global root

def crear_wb_informe():
    wb = openpyxl.Workbook()
    return wb


def unicodeText(text):
    try:
        text = unicodedata.unicode(text, 'utf-8')
    except TypeError:
        return text

def crea_hoja_info(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet

def crea_hoja_info_pdf(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet



def border_tabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

    colfin=colfin+1
    filfin=filfin+2

    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    for i in range(fil, filfin-1):
        for j in range(col, colfin):
            sheet.cell(row=i, column=j).border = border_cell


def columnas_filas(sheet, flag, celda, value):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value



def poner_border(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    sheet.cell(row=fil, column=col).border = border_cell

def Informe(sheet, dic,lista_alumnos,cant_alumno,filtro,filtro1):
    columnas_filas(sheet, 0, 'A', 10.00)
    columnas_filas(sheet, 0, 'B', 5.00)
    columnas_filas(sheet, 0, 'C', 10.00)     
    columnas_filas(sheet, 0, 'D', 7.00)
    columnas_filas(sheet, 0, 'E', 12.00)
    columnas_filas(sheet, 0, 'F', 10.00)
    columnas_filas(sheet, 0, 'G', 10.00)
    columnas_filas(sheet, 0, 'H', 7.00)
    columnas_filas(sheet, 0, 'I', 10.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=6, name='arial')
    fuente3 = Font(bold=True, size=8, name='arial')
    fuente2 = Font(bold=True, size=6, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:I2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente3
    sheet['A2']= 'Listado de Cheques y Ordenes de Pago'

    sheet['H1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['H1'].font = fuente2
    sheet['H1']= 'Usuario'
    usuario_id=str(dic['usuario_id'].encode('utf-8'))
    sheet['I1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I1'].font = fuente
    sheet['I1']= str(dic['usuario_id'].encode('utf-8'))

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente2
    sheet['A1']= 'Compañia'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente
    sheet['B1']= str(dic['company_id'].encode('utf-8'))

    sheet['A4'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A4'].font = fuente2
    sheet['A4']= 'Fecha Emisión:'
    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha_corte']

    sheet.merge_cells('B4:C4')
    sheet['B4'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B4'].font = fuente
    sheet['B4']= fecha_actual

    sheet['A5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A5'].font = fuente2
    sheet['A5']= 'Metodos de Pagos:'

    sheet.merge_cells('B5:E5')
    sheet['B5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B5'].font = fuente
    if filtro==0:
        sheet['B5']= "Todos los metodos de pagos"
    else:
        sheet['B5']= filtro

    sheet['F5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F5'].font = fuente2
    sheet['F5']= 'Estados de Pagos:'

    sheet.merge_cells('G5:G5')
    sheet['G5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G5'].font = fuente
    if filtro1==0:
        sheet['G5']= "Todos"
    else:
        sheet['G5']= filtro1

    poner_border(sheet,1,1,'medium','medium','none','none')
    poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'none','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','none','none')
    poner_border(sheet,1,7,'none','medium','none','none')
    poner_border(sheet,1,8,'none','medium','none','none')
    poner_border(sheet,1,9,'none','medium','medium','none')

    poner_border(sheet,2,1,'medium','none','none','none')
    poner_border(sheet,2,9,'none','none','medium','none')

    poner_border(sheet,3,1,'medium','none','none','medium')
    poner_border(sheet,3,2,'none','none','none','medium')
    poner_border(sheet,3,3,'none','none','none','medium')
    poner_border(sheet,3,4,'none','none','none','medium')
    poner_border(sheet,3,5,'none','none','none','medium')
    poner_border(sheet,3,6,'none','none','none','medium')
    poner_border(sheet,3,7,'none','none','none','medium')
    poner_border(sheet,3,8,'none','none','none','medium')
    poner_border(sheet,3,9,'none','none','medium','medium')
    fecha_ini=dic['fecha_desde']
    fecha_fin=dic['fecha_hasta']
    fecha=str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])
    sheet.merge_cells('D3:F3')
    sheet['D3'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D3'].font = fuente2
    sheet['D3']= str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])

    sheet['A6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A6'].font = fuente2
    sheet['A6']= 'Fecha Emisión'

    sheet.merge_cells('B6:C6')
    sheet['B6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B6'].font = fuente2
    sheet['B6']= 'Egreso #'

    sheet.merge_cells('D6:E6')
    sheet['D6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['D6'].font = fuente2
    sheet['D6']= 'Cheque #'

    sheet['F6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F6'].font = fuente2
    sheet['F6']= 'Beneficiario'

    sheet.merge_cells('G6:H6')
    sheet['G6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G6'].font = fuente2
    sheet['G6']= 'Observación'

    sheet['I6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I6'].font = fuente2
    sheet['I6']= 'Valor'

    fila=7
    total_general=0.0
    saldo_general=0.0
    saldo=0.0
    total=0.0
    dic={}
    lista_datos=[]
    for recorrer in lista_alumnos:
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente
        sheet['A'+str(fila)]= recorrer['fecha_emision']
        #poner_border(sheet,fila,1,'none','none','none','medium')

        sheet.merge_cells('B'+str(fila)+':C'+str(fila))
        sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila)].font = fuente
        if recorrer['egreso']==0:
            sheet['B'+str(fila)]= ""
        else:
            sheet['B'+str(fila)]= recorrer['egreso']

        sheet.merge_cells('D'+str(fila)+':E'+str(fila))
        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila)].font = fuente
        if recorrer['cheque']==0:
            sheet['D'+str(fila)]= ""
        else:
            sheet['D'+str(fila)]= recorrer['cheque']

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila)].font = fuente
        sheet['F'+str(fila)]= recorrer['beneficiario']

        sheet.merge_cells('G'+str(fila)+':H'+str(fila))
        sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['G'+str(fila)].font = fuente
        if recorrer['observacion']==0:
            sheet['G'+str(fila)]= ""
        else:
            sheet['G'+str(fila)]= recorrer['observacion']

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['I'+str(fila)].font = fuente
        sheet['I'+str(fila)]= "{:,}".format(float(recorrer['valor'])).replace(',','~').replace('.',',').replace('~','.')

        
        total_general=total_general+float(recorrer['valor'])
        fila=fila+1

    sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G'+str(fila)].font = fuente2
    sheet['G'+str(fila)]= 'TOTAL'

    sheet.merge_cells('H'+str(fila)+':I'+str(fila))
    sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['H'+str(fila)].font = fuente2
    sheet['H'+str(fila)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')

    columnas_filas(sheet, 1, str(fila+5), 10.00)
    sheet.merge_cells('D'+str(fila+5)+':F'+str(fila+5))
    sheet['D'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['D'+str(fila+5)].font = fuente2
    sheet['D'+str(fila+5)]= usuario_id
    poner_border(sheet,fila+5,4,'none','thin','none','none')
    poner_border(sheet,fila+5,5,'none','thin','none','none')
    poner_border(sheet,fila+5,6,'none','thin','none','none')

    


def Informe_pdf(sheet, dic,lista_alumnos,cant_alumno,filtro,filtro1):
    columnas_filas(sheet, 0, 'A', 10.00)
    columnas_filas(sheet, 0, 'B', 5.00)
    columnas_filas(sheet, 0, 'C', 10.00)     
    columnas_filas(sheet, 0, 'D', 7.00)
    columnas_filas(sheet, 0, 'E', 1.00)
    columnas_filas(sheet, 0, 'F', 12.00)
    columnas_filas(sheet, 0, 'G', 10.00)
    columnas_filas(sheet, 0, 'H', 10.00)
    columnas_filas(sheet, 0, 'I', 7.00)
    columnas_filas(sheet, 0, 'J', 10.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=6, name='arial')
    fuente3 = Font(bold=True, size=8, name='arial')
    fuente2 = Font(bold=True, size=6, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('A2:I2')
    sheet['A2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['A2'].font = fuente3
    sheet['A2']= 'Listado de Cheques y Ordenes de Pago'

    sheet['H1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['H1'].font = fuente2
    sheet['H1']= 'Usuario'
    usuario_id=str(dic['usuario_id'].encode('utf-8'))
    sheet['I1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I1'].font = fuente
    sheet['I1']= str(dic['usuario_id'].encode('utf-8'))

    sheet['A1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A1'].font = fuente2
    sheet['A1']= 'Compañia'

    sheet.merge_cells('B1:C1')
    sheet['B1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B1'].font = fuente
    sheet['B1']= str(dic['company_id'].encode('utf-8'))

    sheet['A4'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A4'].font = fuente2
    sheet['A4']= 'Fecha Emisión:'
    #fecha_actual = datetime.strftime(datetime.now(), '%d-%m-%Y %H:%M:%S')
    fecha_actual = dic['fecha_corte']

    sheet.merge_cells('B4:C4')
    sheet['B4'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B4'].font = fuente
    sheet['B4']= fecha_actual

    sheet['A5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A5'].font = fuente2
    sheet['A5']= 'Metodos de Pagos:'

    sheet.merge_cells('B5:E5')
    sheet['B5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B5'].font = fuente
    if filtro==0:
        sheet['B5']= "Todos los metodos de pagos"
    else:
        sheet['B5']= filtro

    sheet['F5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F5'].font = fuente2
    sheet['F5']= 'Estados de Pagos:'

    sheet.merge_cells('G5:G5')
    sheet['G5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G5'].font = fuente
    if filtro1==0:
        sheet['G5']= "Todos"
    else:
        sheet['G5']= filtro1

    poner_border(sheet,1,1,'medium','medium','none','none')
    poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'none','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','none','none')
    poner_border(sheet,1,7,'none','medium','none','none')
    poner_border(sheet,1,8,'none','medium','none','none')
    poner_border(sheet,1,9,'none','medium','medium','none')

    poner_border(sheet,2,1,'medium','none','none','none')
    poner_border(sheet,2,9,'none','none','medium','none')

    poner_border(sheet,3,1,'medium','none','none','medium')
    poner_border(sheet,3,2,'none','none','none','medium')
    poner_border(sheet,3,3,'none','none','none','medium')
    poner_border(sheet,3,4,'none','none','none','medium')
    poner_border(sheet,3,5,'none','none','none','medium')
    poner_border(sheet,3,6,'none','none','none','medium')
    poner_border(sheet,3,7,'none','none','none','medium')
    poner_border(sheet,3,8,'none','none','none','medium')
    poner_border(sheet,3,9,'none','none','medium','medium')
    fecha_ini=dic['fecha_desde']
    fecha_fin=dic['fecha_hasta']
    fecha=str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])
    sheet.merge_cells('D3:F3')
    sheet['D3'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D3'].font = fuente2
    sheet['D3']= str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])

    sheet['A6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A6'].font = fuente2
    sheet['A6']= 'Fecha Emisión'

    sheet.merge_cells('B6:C6')
    sheet['B6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B6'].font = fuente2
    sheet['B6']= 'Egreso #'

    sheet.merge_cells('D6:E6')
    sheet['D6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['D6'].font = fuente2
    sheet['D6']= 'Cheque #'

    sheet['F6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F6'].font = fuente2
    sheet['F6']= 'Beneficiario'

    sheet.merge_cells('G6:H6')
    sheet['G6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G6'].font = fuente2
    sheet['G6']= 'Observación'

    sheet['I6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I6'].font = fuente2
    sheet['I6']= 'Valor'

    fila=7
    total_general=0.0
    saldo_general=0.0
    saldo=0.0
    total=0.0
    dic={}
    lista_datos=[]
    for recorrer in lista_alumnos:
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente
        sheet['A'+str(fila)]= recorrer['fecha_emision']
        #poner_border(sheet,fila,1,'none','none','none','medium')

        sheet.merge_cells('B'+str(fila)+':C'+str(fila))
        sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila)].font = fuente
        if recorrer['egreso']==0:
            sheet['B'+str(fila)]= ""
        else:
            sheet['B'+str(fila)]= recorrer['egreso']

        sheet.merge_cells('D'+str(fila)+':E'+str(fila))
        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila)].font = fuente
        if recorrer['cheque']==0:
            sheet['D'+str(fila)]= ""
        else:
            sheet['D'+str(fila)]= recorrer['cheque']

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila)].font = fuente
        sheet['F'+str(fila)]= recorrer['beneficiario']

        sheet.merge_cells('G'+str(fila)+':H'+str(fila))
        sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['G'+str(fila)].font = fuente
        if recorrer['observacion']==0:
            sheet['G'+str(fila)]= ""
        else:
            sheet['G'+str(fila)]= recorrer['observacion']

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['I'+str(fila)].font = fuente
        sheet['I'+str(fila)]= "{:,}".format(float(recorrer['valor'])).replace(',','~').replace('.',',').replace('~','.')

        
        total_general=total_general+float(recorrer['valor'])
        fila=fila+1

    sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G'+str(fila)].font = fuente2
    sheet['G'+str(fila)]= 'TOTAL'

    sheet.merge_cells('H'+str(fila)+':I'+str(fila))
    sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['H'+str(fila)].font = fuente2
    sheet['H'+str(fila)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')

    columnas_filas(sheet, 1, str(fila+5), 10.00)
    sheet.merge_cells('D'+str(fila+5)+':F'+str(fila+5))
    sheet['D'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['D'+str(fila+5)].font = fuente2
    sheet['D'+str(fila+5)]= usuario_id
    poner_border(sheet,fila+5,4,'none','thin','none','none')
    poner_border(sheet,fila+5,5,'none','thin','none','none')
    poner_border(sheet,fila+5,6,'none','thin','none','none')
