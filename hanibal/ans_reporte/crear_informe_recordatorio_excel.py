# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import openpyxl.worksheet
import unicodedata
from copy import deepcopy
from openpyxl.chart import (
    Reference,
    Series,
    BarChart
)
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime, date, timedelta
import time
import locale
global root

def crear_wb_informe():
    wb = openpyxl.Workbook()
    return wb


def unicodeText(text):
    try:
        text = unicodedata.unicode(text, 'utf-8')
    except TypeError:
        return text

def crea_hoja_info(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        #sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet

def crea_hoja_info_pdf(wb, title, flag):
    sheet = wb.active
    if(flag == 0):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        #sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = False
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.print_options.horizontalCentered = True
    if(flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_A4_SMALL
        #sheet.print_options.scale = 100
        #sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = False
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.print_options.horizontalCentered = True
    sheet.title = title
    return sheet



def border_tabla(sheet, col, colfin, fil, filfin, styleleft, styletop, styleright, stylebottom):

    colfin=colfin+1
    filfin=filfin+2

    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    for i in range(fil, filfin-1):
        for j in range(col, colfin):
            sheet.cell(row=i, column=j).border = border_cell


def columnas_filas(sheet, flag, celda, value):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value



def poner_border(sheet, fil, col, styleleft, styletop, styleright, stylebottom):
    border_cell = Border(left=Side(style=styleleft), top=Side(style=styletop), right=Side(style=styleright), bottom=Side(style=stylebottom))
    sheet.cell(row=fil, column=col).border = border_cell

def Informe(sheet, dic,lista_alumnos):
    columnas_filas(sheet, 0, 'A', 3.00)
    columnas_filas(sheet, 0, 'B', 4.00)
    columnas_filas(sheet, 0, 'C', 2.00)     
    columnas_filas(sheet, 0, 'D', 6.00)
    columnas_filas(sheet, 0, 'E', 12.00)
    columnas_filas(sheet, 0, 'F', 9.00)
    columnas_filas(sheet, 0, 'G', 6.00)
    columnas_filas(sheet, 0, 'H', 16.00)
    columnas_filas(sheet, 0, 'I', 16.00)
    columnas_filas(sheet, 0, 'J', 12.00)
    columnas_filas(sheet, 0, 'K', 9.00)
    columnas_filas(sheet, 0, 'L', 6.00)
    columnas_filas(sheet, 0, 'M', 6.00)
    columnas_filas(sheet, 0, 'N', 5.00)
    columnas_filas(sheet, 0, 'O', 5.00)
    columnas_filas(sheet, 0, 'P', 5.00)
    columnas_filas(sheet, 0, 'Q', 7.00)
    columnas_filas(sheet, 0, 'R', 5.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=5, name='arial')
    fuente3 = Font(bold=True, size=7, name='arial')
    fuente2 = Font(bold=True, size=5, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('D2:N2')
    sheet['D2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D2'].font = fuente3
    sheet['D2']= 'Detalle de Recordatorios de Pago'

    #sheet.merge_cells('L1:M1')
    sheet['L1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['L1'].font = fuente2
    sheet['L1']= 'Usuario'

    usuario_id=str(dic['usuario_id'].encode('utf-8'))
    sheet.merge_cells('M1:N1')
    sheet['M1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['M1'].font = fuente
    sheet['M1']= str(dic['usuario_id'].encode('utf-8'))

    sheet.merge_cells('C1:F1')
    sheet['C1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C1'].font = fuente2
    sheet['C1']= 'Compañia: '+str(dic['company_id'].encode('utf-8'))


    #poner_border(sheet,1,1,'medium','medium','none','none')
    #poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'medium','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','none','none')
    poner_border(sheet,1,7,'none','medium','none','none')
    poner_border(sheet,1,8,'none','medium','none','none')
    poner_border(sheet,1,9,'none','medium','none','none')
    poner_border(sheet,1,10,'none','medium','none','none')
    poner_border(sheet,1,11,'none','medium','none','none')
    poner_border(sheet,1,12,'none','medium','none','none')
    poner_border(sheet,1,13,'none','medium','none','none')
    poner_border(sheet,1,14,'none','medium','none','none')
    poner_border(sheet,1,15,'none','medium','medium','none')

    poner_border(sheet,2,3,'medium','none','none','none')
    poner_border(sheet,2,15,'none','none','medium','none')
    poner_border(sheet,3,3,'medium','none','none','none')
    poner_border(sheet,3,15,'none','none','medium','none')


    poner_border(sheet,4,3,'none','medium','none','none')
    poner_border(sheet,4,4,'none','medium','none','none')
    poner_border(sheet,4,5,'none','medium','none','none')
    poner_border(sheet,4,6,'none','medium','none','none')
    poner_border(sheet,4,7,'none','medium','none','none')
    poner_border(sheet,4,8,'none','medium','none','none')
    poner_border(sheet,4,9,'none','medium','none','none')
    poner_border(sheet,4,10,'none','medium','none','none')
    poner_border(sheet,4,11,'none','medium','none','none')
    poner_border(sheet,4,12,'none','medium','none','none')
    poner_border(sheet,4,13,'none','medium','none','none')
    poner_border(sheet,4,14,'none','medium','none','none')
    poner_border(sheet,4,15,'none','medium','none','none')

    fecha_ini=dic['fecha_desde']
    fecha_fin=dic['fecha_hasta']
    fecha=str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])
    sheet.merge_cells('D3:N3')
    sheet['D3'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D3'].font = fuente2
    sheet['D3']= str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])

    sheet.merge_cells('A5:F5')
    sheet['A5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A5'].font = fuente2
    sheet['A5']= str('Fecha Emisión: '+str(dic['fecha_emision']))

    sheet.merge_cells('A6:C6')
    sheet['A6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A6'].font = fuente2
    jornada = 'Jornada: %s' % (dic['jornada'] if dic['jornada'] else '')
    sheet['A6']= jornada

    sheet.merge_cells('F6:G6')
    sheet['F6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F6'].font = fuente2
    seccion = 'Seccion: %s' % (dic['seccion'] if dic['seccion'] else '')
    sheet['F6']= seccion

    sheet.merge_cells('H6:I6')
    sheet['H6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H6'].font = fuente2
    curso = 'Curso: %s' % (dic['curso'] if dic['curso'] else '')
    sheet['H6']= curso

    sheet.merge_cells('M6:N6')
    sheet['M6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['M6'].font = fuente2
    paralelo = 'Paralelo: %s' % (dic['paralelo'] if dic['paralelo'] else '')
    sheet['M6']= paralelo

    sheet.merge_cells('A7:F7')
    sheet['A7'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A7'].font = fuente2
    representante = 'Representante: %s' % (dic['representante'] if dic['representante'] else '')
    sheet['A7']= representante

    sheet.merge_cells('H7:M7')
    sheet['H7'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H7'].font = fuente2
    alumno = 'Alumno: %s'% (dic['alumno'] if dic['alumno'] else '')
    sheet['H7']= alumno

    fecha_actual = dic['fecha_emision']

    sheet['A8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A8'].font = fuente2
    sheet['A8']= 'Orden'

    sheet.merge_cells('B8:C8')
    sheet['B8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B8'].font = fuente2
    sheet['B8']= 'Codigo'

    sheet['D8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['D8'].font = fuente2
    sheet['D8']= 'Jornada'

    sheet['E8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['E8'].font = fuente2
    sheet['E8']= 'Seccion'

    sheet['F8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F8'].font = fuente2
    sheet['F8']= 'Curso'

    sheet['G8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G8'].font = fuente2
    sheet['G8']= 'Paralelo'

    sheet['H8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H8'].font = fuente2
    sheet['H8']= 'Alumno'

    sheet['I8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I8'].font = fuente2
    sheet['I8']= 'Representante'

    sheet['J8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['J8'].font = fuente2
    sheet['J8']= 'Correo'

    sheet['K8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['K8'].font = fuente2
    sheet['K8']= '# Factura'

    sheet['L8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['L8'].font = fuente2
    sheet['L8']= 'Fecha Factura'

    sheet.merge_cells('M8:N8')
    sheet['M8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['M8'].font = fuente2
    sheet['M8']= 'Concepto'

    sheet['O8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['O8'].font = fuente2
    sheet['O8']= 'Monto'

    sheet['P8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['P8'].font = fuente2
    sheet['P8']= 'Saldo'

    sheet['Q8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['Q8'].font = fuente2
    sheet['Q8']= '# Notificacion'

    sheet['R8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['R8'].font = fuente2
    sheet['R8']= 'Fecha envio'

    fila=9
    total_general=0.0
    saldo_general=0.0
    saldo=0.0
    total=0.0
    dic={}
    lista_datos=[]
    for recorrer in lista_alumnos:
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente
        sheet['A'+str(fila)]= recorrer['orden']
        #poner_border(sheet,fila,1,'none','none','none','medium')

        sheet.merge_cells('B'+str(fila)+':C'+str(fila))
        sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila)].font = fuente
        sheet['B'+str(fila)]= recorrer['sequence']

        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila)].font = fuente
        sheet['D'+str(fila)]= recorrer['jornada']

        sheet['E'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['E'+str(fila)].font = fuente
        sheet['E'+str(fila)]= recorrer['seccion']

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila)].font = fuente
        sheet['F'+str(fila)]= recorrer['curso']

        sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['G'+str(fila)].font = fuente
        sheet['G'+str(fila)]= recorrer['paralelo']

        sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['H'+str(fila)].font = fuente
        sheet['H'+str(fila)]= recorrer['alumno']

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['I'+str(fila)].font = fuente
        sheet['I'+str(fila)]= recorrer['representante']

        sheet['J'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['J'+str(fila)].font = fuente
        sheet['J'+str(fila)]= recorrer['correo']

        sheet['K'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['K'+str(fila)].font = fuente
        sheet['K'+str(fila)]= recorrer['numero_factura']

        sheet['L'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['L'+str(fila)].font = fuente
        sheet['L'+str(fila)]= recorrer['fecha_factura']

        sheet.merge_cells('M'+str(fila)+':N'+str(fila))
        sheet['M'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
        sheet['M'+str(fila)].font = fuente
        sheet['M'+str(fila)]= recorrer['concepto']

        sheet['O'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['O'+str(fila)].font = fuente
        sheet['O'+str(fila)]= "{:,}".format(float(recorrer['monto'])).replace(',','~').replace('.',',').replace('~','.')

        sheet['P'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['P'+str(fila)].font = fuente
        sheet['P'+str(fila)]= "{:,}".format(float(recorrer['saldo'])).replace(',','~').replace('.',',').replace('~','.')

        sheet['Q'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['Q'+str(fila)].font = fuente
        sheet['Q'+str(fila)]= recorrer['notificaciones']

        sheet['R'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['R'+str(fila)].font = fuente
        sheet['R'+str(fila)]= recorrer['fecha_envio_correo']  
        
        total_general=total_general+float(recorrer['monto'])
        saldo_general=saldo_general+float(recorrer['saldo'])
        fila=fila+1

    sheet.merge_cells('M'+str(fila)+':N'+str(fila))
    sheet['M'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['M'+str(fila)].font = fuente2
    sheet['M'+str(fila)]= 'TOTAL'

    
    sheet['O'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['O'+str(fila)].font = fuente2
    sheet['O'+str(fila)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')

    sheet['P'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['P'+str(fila)].font = fuente2
    sheet['P'+str(fila)]= "{:,}".format(float(saldo_general)).replace(',','~').replace('.',',').replace('~','.')

    columnas_filas(sheet, 1, str(fila+5), 10.00)
    sheet.merge_cells('H'+str(fila+5)+':I'+str(fila+5))
    sheet['H'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['H'+str(fila+5)].font = fuente2
    sheet['H'+str(fila+5)]= usuario_id
    poner_border(sheet,fila+5,8,'none','thin','none','none')
    poner_border(sheet,fila+5,9,'none','thin','none','none')


    


def Informe_pdf(sheet, dic,lista_alumnos):
    columnas_filas(sheet, 0, 'A', 4.00)
    columnas_filas(sheet, 0, 'B', 4.00)
    columnas_filas(sheet, 0, 'C', 2.00)     
    columnas_filas(sheet, 0, 'D', 10.00)
    columnas_filas(sheet, 0, 'E', 4.00)
    columnas_filas(sheet, 0, 'F', 14.00)
    columnas_filas(sheet, 0, 'G', 14.00)
    columnas_filas(sheet, 0, 'H', 12.00)
    columnas_filas(sheet, 0, 'I', 7.00)
    columnas_filas(sheet, 0, 'J', 5.00)
    columnas_filas(sheet, 0, 'K', 10.00)
    columnas_filas(sheet, 0, 'L', 3.00)
    columnas_filas(sheet, 0, 'M', 4.00)
    columnas_filas(sheet, 0, 'N', 4.00)
    columnas_filas(sheet, 0, 'O', 5.00)
    columnas_filas(sheet, 0, 'P', 5.00)

    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente = Font(bold=False, size=5, name='arial')
    fuente3 = Font(bold=True, size=10, name='arial')
    fuente2 = Font(bold=True, size=5, name='arial')

    fila = 3
    fila1 = 2
    acum=1
    cont=0
    col=2
    col1=4
    fil=4
    coli=2
    colf=2

    sheet.merge_cells('D2:N2')
    sheet['D2'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D2'].font = fuente3
    sheet['D2']= 'Detalle de Recordatorios de Pago'

    #sheet.merge_cells('L1:M1')
    sheet['L1'].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['L1'].font = fuente2
    sheet['L1']= 'Usuario'

    usuario_id=str(dic['usuario_id'].encode('utf-8'))
    sheet.merge_cells('M1:N1')
    sheet['M1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['M1'].font = fuente
    sheet['M1']= str(dic['usuario_id'].encode('utf-8'))

    sheet.merge_cells('C1:F1')
    sheet['C1'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['C1'].font = fuente2
    sheet['C1']= 'Compañia: '+str(dic['company_id'].encode('utf-8'))


    #poner_border(sheet,1,1,'medium','medium','none','none')
    #poner_border(sheet,1,2,'none','medium','none','none')
    poner_border(sheet,1,3,'medium','medium','none','none')
    poner_border(sheet,1,4,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,5,'none','medium','none','none')
    poner_border(sheet,1,6,'none','medium','none','none')
    poner_border(sheet,1,7,'none','medium','none','none')
    poner_border(sheet,1,8,'none','medium','none','none')
    poner_border(sheet,1,9,'none','medium','none','none')
    poner_border(sheet,1,10,'none','medium','none','none')
    poner_border(sheet,1,11,'none','medium','none','none')
    poner_border(sheet,1,12,'none','medium','none','none')
    poner_border(sheet,1,13,'none','medium','none','none')
    poner_border(sheet,1,14,'none','medium','none','none')
    poner_border(sheet,1,15,'none','medium','medium','none')

    poner_border(sheet,2,3,'medium','none','none','none')
    poner_border(sheet,2,15,'none','none','medium','none')
    poner_border(sheet,3,3,'medium','none','none','none')
    poner_border(sheet,3,15,'none','none','medium','none')


    poner_border(sheet,4,3,'none','medium','none','none')
    poner_border(sheet,4,4,'none','medium','none','none')
    poner_border(sheet,4,5,'none','medium','none','none')
    poner_border(sheet,4,6,'none','medium','none','none')
    poner_border(sheet,4,7,'none','medium','none','none')
    poner_border(sheet,4,8,'none','medium','none','none')
    poner_border(sheet,4,9,'none','medium','none','none')
    poner_border(sheet,4,10,'none','medium','none','none')
    poner_border(sheet,4,11,'none','medium','none','none')
    poner_border(sheet,4,12,'none','medium','none','none')
    poner_border(sheet,4,13,'none','medium','none','none')
    poner_border(sheet,4,14,'none','medium','none','none')
    poner_border(sheet,4,15,'none','medium','none','none')

    fecha_ini=dic['fecha_desde']
    fecha_fin=dic['fecha_hasta']
    fecha=str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])
    sheet.merge_cells('D3:N3')
    sheet['D3'].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
    sheet['D3'].font = fuente2
    sheet['D3']= str(" Desde: "+dic['fecha_desde']+"    Hasta: "+dic['fecha_hasta'])

    sheet.merge_cells('A5:F5')
    sheet['A5'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A5'].font = fuente2
    sheet['A5']= str('Fecha Emisión: '+str(dic['fecha_emision']))

    sheet.merge_cells('A6:C6')
    sheet['A6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A6'].font = fuente2
    jornada = 'Jornada: %s' % (dic['jornada'] if dic['jornada'] else '')
    sheet['A6']= jornada

    sheet.merge_cells('F6:G6')
    sheet['F6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F6'].font = fuente2
    seccion = 'Seccion: %s' % (dic['seccion'] if dic['seccion'] else '')
    sheet['F6']= seccion

    sheet.merge_cells('H6:I6')
    sheet['H6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H6'].font = fuente2
    curso = 'Curso: %s' % (dic['curso'] if dic['curso'] else '')
    sheet['H6']= curso

    sheet.merge_cells('M6:N6')
    sheet['M6'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['M6'].font = fuente2
    paralelo = 'Paralelo: %s' % (dic['paralelo'] if dic['paralelo'] else '')
    sheet['M6']= paralelo

    sheet.merge_cells('A7:F7')
    sheet['A7'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A7'].font = fuente2
    representante = 'Representante: %s' % (dic['representante'] if dic['representante'] else '')
    sheet['A7']= representante

    sheet.merge_cells('H7:M7')
    sheet['H7'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H7'].font = fuente2
    alumno = 'Alumno: %s'% (dic['alumno'] if dic['alumno'] else '')
    sheet['H7']= alumno

    fecha_actual = dic['fecha_emision']

    sheet['A8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['A8'].font = fuente2
    sheet['A8']= 'Orden'

    sheet.merge_cells('B8:C8')
    sheet['B8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['B8'].font = fuente2
    sheet['B8']= 'Codigo'

    sheet.merge_cells('D8:E8')
    sheet['D8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['D8'].font = fuente2
    sheet['D8']= 'Estructura Escolar'

    sheet['F8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['F8'].font = fuente2
    sheet['F8']= 'Alumno'

    sheet['G8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['G8'].font = fuente2
    sheet['G8']= 'Representante'

    sheet['H8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['H8'].font = fuente2
    sheet['H8']= 'Correo'

    sheet['I8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['I8'].font = fuente2
    sheet['I8']= '# Factura'

    sheet['J8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['J8'].font = fuente2
    sheet['J8']= 'Fecha Factura'

    sheet.merge_cells('K8:L8')
    sheet['K8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['K8'].font = fuente2
    sheet['K8']= 'Concepto'

    sheet['M8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['M8'].font = fuente2
    sheet['M8']= 'Monto'

    sheet['N8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['N8'].font = fuente2
    sheet['N8']= 'Saldo'

    sheet['O8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['O8'].font = fuente2
    sheet['O8']= '# Notificacion'

    sheet['P8'].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['P8'].font = fuente2
    sheet['P8']= 'Fecha envio'

    fila=9
    total_general=0.0
    saldo_general=0.0
    saldo=0.0
    total=0.0
    dic={}
    lista_datos=[]
    for recorrer in lista_alumnos:
        sheet['A'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['A'+str(fila)].font = fuente
        sheet['A'+str(fila)]= recorrer['orden']
        #poner_border(sheet,fila,1,'none','none','none','medium')

        sheet.merge_cells('B'+str(fila)+':C'+str(fila))
        sheet['B'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['B'+str(fila)].font = fuente
        sheet['B'+str(fila)]= recorrer['sequence']

        sheet.merge_cells('D'+str(fila)+':E'+str(fila))
        sheet['D'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['D'+str(fila)].font = fuente
        sheet['D'+str(fila)]= recorrer['descripcion']

        sheet['F'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['F'+str(fila)].font = fuente
        sheet['F'+str(fila)]= recorrer['alumno']

        sheet['G'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['G'+str(fila)].font = fuente
        sheet['G'+str(fila)]= recorrer['representante']

        sheet['H'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['H'+str(fila)].font = fuente
        sheet['H'+str(fila)]= recorrer['correo']

        sheet['I'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['I'+str(fila)].font = fuente
        sheet['I'+str(fila)]= recorrer['numero_factura']

        sheet['J'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
        sheet['J'+str(fila)].font = fuente
        sheet['J'+str(fila)]= recorrer['fecha_factura']

        sheet.merge_cells('K'+str(fila)+':L'+str(fila))
        sheet['K'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='justify', vertical='top')
        sheet['K'+str(fila)].font = fuente
        sheet['K'+str(fila)]= recorrer['concepto']

        sheet['M'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['M'+str(fila)].font = fuente
        sheet['M'+str(fila)]= "{:,}".format(float(recorrer['monto'])).replace(',','~').replace('.',',').replace('~','.')

        sheet['N'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
        sheet['N'+str(fila)].font = fuente
        sheet['N'+str(fila)]= "{:,}".format(float(recorrer['saldo'])).replace(',','~').replace('.',',').replace('~','.')

        sheet['O'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['O'+str(fila)].font = fuente
        sheet['O'+str(fila)]= recorrer['notificaciones']

        sheet['P'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='top')
        sheet['P'+str(fila)].font = fuente
        sheet['P'+str(fila)]= recorrer['fecha_envio_correo']
        
        total_general=total_general+float(recorrer['monto'])
        saldo_general=saldo_general+float(recorrer['saldo'])
        fila=fila+1

    sheet.merge_cells('K'+str(fila)+':L'+str(fila))
    sheet['K'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='left', vertical='top')
    sheet['K'+str(fila)].font = fuente2
    sheet['K'+str(fila)]= 'TOTAL'

    
    sheet['M'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['M'+str(fila)].font = fuente2
    sheet['M'+str(fila)]= "{:,}".format(float(total_general)).replace(',','~').replace('.',',').replace('~','.')

    sheet['N'+str(fila)].alignment = alignment_title.copy(wrapText=True,horizontal='right', vertical='top')
    sheet['N'+str(fila)].font = fuente2
    sheet['N'+str(fila)]= "{:,}".format(float(saldo_general)).replace(',','~').replace('.',',').replace('~','.')

    columnas_filas(sheet, 1, str(fila+5), 10.00)
    sheet.merge_cells('H'+str(fila+5)+':I'+str(fila+5))
    sheet['H'+str(fila+5)].alignment = alignment_title.copy(wrapText=True,horizontal='center', vertical='center')
    sheet['H'+str(fila+5)].font = fuente2
    sheet['H'+str(fila+5)]= usuario_id
    poner_border(sheet,fila+5,8,'none','thin','none','none')
    poner_border(sheet,fila+5,9,'none','thin','none','none')
