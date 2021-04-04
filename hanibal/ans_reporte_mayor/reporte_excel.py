#-*- coding: utf-8 -*-
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.worksheet
import openpyxl

import datetime
from datetime import datetime

import logging
from openerp.osv import fields
_logger = logging.getLogger(__name__)
thin_border = Border(bottom=Side(style='hair'))
thin_border_left = Border(left=Side(style='hair'))
thin_border_top = Border(top=Side(style='hair'))
thin_border_right = Border(right=Side(style='hair'))
thin_border_left_top = Border(left=Side(style='hair'),top=Side(style='hair'))
thin_border_top_right = Border(top=Side(style='hair'),right=Side(style='hair'))
thin_border_buttonm_right = Border(right=Side(style='hair'),bottom=Side(style='hair'))
thin_border_buttonm_left = Border(left=Side(style='hair'),bottom=Side(style='hair'))
def ajustes_hoja(sheet, flag, celda, value, value2):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value2
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value

def crear_hoja_libro(wb,title,flag,hojas):
    if(flag == 0):
        sheet = wb.active
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = hojas
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
    if(flag==1):
        sheet = wb.create_sheet()
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = hojas
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

    sheet.title=title
    return sheet

def crear_encabezado(data,sheet_libromayor,alignment_title,fuente_cabecera):
    sheet_libromayor['A1'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['A1'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['A1'] = 'Compañia'


    sheet_libromayor['B1'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['B1'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B1'] = data['plan_cuenta'].name

    sheet_libromayor['B1'].border = thin_border_top

    sheet_libromayor['A1'].border = thin_border_left_top
    sheet_libromayor['A2'].border = thin_border_left
    sheet_libromayor['A3'].border = thin_border_left
    sheet_libromayor['A4'].border = thin_border_left
    sheet_libromayor['A4'].border = thin_border_left
    sheet_libromayor['A5'].border = thin_border_buttonm_left

    sheet_libromayor['K1'].border = thin_border_top_right
    sheet_libromayor['K2'].border = thin_border_right
    sheet_libromayor['K3'].border = thin_border_right
    sheet_libromayor['K4'].border = thin_border_right
    sheet_libromayor['K4'].border = thin_border_right
    sheet_libromayor['K5'].border = thin_border_buttonm_right

    sheet_libromayor['B5'].border = thin_border
    sheet_libromayor['C5'].border = thin_border
    sheet_libromayor['D5'].border = thin_border
    sheet_libromayor['E5'].border = thin_border
    sheet_libromayor['F5'].border = thin_border
    sheet_libromayor['G5'].border = thin_border
    sheet_libromayor['H5'].border = thin_border
    sheet_libromayor['I5'].border = thin_border
    sheet_libromayor['J5'].border = thin_border

    sheet_libromayor['B1'].border = thin_border_top
    sheet_libromayor['C1'].border = thin_border_top
    sheet_libromayor['D1'].border = thin_border_top
    sheet_libromayor['E1'].border = thin_border_top
    sheet_libromayor['F1'].border = thin_border_top
    sheet_libromayor['G1'].border = thin_border_top
    sheet_libromayor['H1'].border = thin_border_top
    sheet_libromayor['I1'].border = thin_border_top
    sheet_libromayor['J1'].border = thin_border_top
    sheet_libromayor['K1'].border = thin_border_top_right

    sheet_libromayor['J1'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['J1'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['J1'] = 'Usuario:'

    sheet_libromayor['H1'].border = thin_border_top
    sheet_libromayor['C1'].border = thin_border_top
    sheet_libromayor['D1'].border = thin_border_top
    sheet_libromayor['E1'].border = thin_border_top
    sheet_libromayor['F1'].border = thin_border_top
    sheet_libromayor['G1'].border = thin_border_top
    sheet_libromayor['I1'].border = thin_border_top
    sheet_libromayor['J1'].border = thin_border_top
    sheet_libromayor['K1'].border = thin_border_top



    sheet_libromayor['K1'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['K1'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['K1'] = data['usuario']

    sheet_libromayor['A2'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['A2'].font = fuente_cabecera.copy(bold=True, size=20, name="calibri", color='000000')
    sheet_libromayor['A2'] = 'Mayor Auxiliar'
    sheet_libromayor.row_dimensions[2].height = float(31.50)

    sheet_libromayor.merge_cells('A2:K2')

    sheet_libromayor['D3'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['D3'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['D3'] = 'Desde: '

    sheet_libromayor['E3'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['E3'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['E3'] = data['fecha_desde']

    sheet_libromayor['F3'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['F3'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['F3'] = 'Hasta: '

    sheet_libromayor['G3'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['G3'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['G3'] = data['fecha_hasta']

    sheet_libromayor['A5'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['A5'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['A5'] = 'Plan de cuenta'

    sheet_libromayor['B5'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['B5'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B5'] = data['plan_cuenta'].name

    sheet_libromayor['I5'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                         vertical="center")
    sheet_libromayor['I5'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['I5'] = 'Fecha Emision: '

    sheet_libromayor['J5'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['J5'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['J5'] = fields.datetime.now()
    sheet_libromayor.merge_cells('J5:K5')





    sheet_libromayor['H6'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['H6'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['H6'] = 'Saldo Total Inicial: '

    sheet_libromayor.merge_cells('H6:I6')
    sheet_libromayor['J6'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['J6'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['J6'] = data['inicial']
    sheet_libromayor['J6'].number_format = '0.00'

    size_celdas(20.00, sheet_libromayor)


def crear_reporte_excell():
    libro = openpyxl.Workbook(encoding="UTF-8", data_only=True)
    sheet_libromayor = crear_hoja_libro(libro, 'Reporte Libro Mayor', 0, 0)
    sheet_view = openpyxl.worksheet.SheetView()
    sheet_view.zoomScale = "85"
    sheet_view.zoomScaleNormal = "70"
    sheet_libromayor.sheet_view = sheet_view
    sheet_libromayor.zoomScale = "70"
    ajustes_hoja(sheet_libromayor, 0, 'A', 20.00, 5.00)
    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente_cabecera = Font(bold=True, size=15, name='calibri')

    all_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    dic = {'fuente':fuente_cabecera,'border':all_border,'alineacion':alignment_title,'sheet':sheet_libromayor,'libro':libro}

    return dic

def mes_actual(date):
    meses = {'January':'Enero','February':'Febrero','March':'Marzo','April':'Abril','May':'Mayo','June':'Junio'
        ,'July':'Julio','August':'Agosto','September':'Septiembe','October':'Octubre','November':'Noviembre','December':'Diciembre'}
    return meses[date]

def cuerpo_reporte(sheet_libromayor,alignment_title,fuente_cabecera,data,hoja,hojas,hoja_cuerpo):
    sheet_libromayor['A'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['A'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['A'+str(hoja)] = 'Cuenta: '
    sheet_libromayor['A' + str(hoja)].border = thin_border

    sheet_libromayor['B'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['B'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['B'+str(hoja)] = str(data['cuenta'])
    sheet_libromayor['B' + str(hoja)].border = thin_border
    sheet_libromayor['C' + str(hoja)].border = thin_border

    sheet_libromayor.merge_cells('B'+str(hoja)+':C'+str(hoja))

    sheet_libromayor['F'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['F'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['F'+str(hoja)] = 'Saldo Inicial: '
    sheet_libromayor['F' + str(hoja)].border = thin_border
    sheet_libromayor['G' + str(hoja)].border = thin_border
    #sheet_libromayor.merge_cells('F' + str(hoja) + ':G' + str(hoja))

    sheet_libromayor['G'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                            vertical="center")
    sheet_libromayor['G'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['G'+str(hoja)] = data['ultimo_saldo']
    sheet_libromayor['G' + str(hoja)].number_format = '0.00'
    sheet_libromayor['H'+str(hoja)].border = thin_border


    sheet_libromayor['I'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                            vertical="center")
    sheet_libromayor['I'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['I'+str(hoja)] = data['fecha']
    sheet_libromayor['I'+str(hoja)].border = thin_border

    sheet_libromayor['A8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['A8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['A8'] = 'ASIENTOS'
    sheet_libromayor['A8'].border = thin_border

    sheet_libromayor['B8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['B8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B8'] = 'DOCUMENTO'
    sheet_libromayor['B8'].border = thin_border

    sheet_libromayor['C8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['C8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['C8'] = 'FECHA'
    sheet_libromayor['C8'].border = thin_border

    sheet_libromayor['D8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['D8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['D8'] = 'DETALLE'
    sheet_libromayor['D8'].border = thin_border

    sheet_libromayor['E8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['E8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['E8'] = 'ALUMNO'
    sheet_libromayor['E8'].border = thin_border

    sheet_libromayor['F8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['F8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F8'] = 'JORNADA'
    sheet_libromayor['F8'].border = thin_border

    sheet_libromayor['G8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['G8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['G8'] = 'SECCION'
    sheet_libromayor['G8'].border = thin_border

    sheet_libromayor['H8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['H8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['H8'] = 'CURSO'
    sheet_libromayor['H8'].border = thin_border

    sheet_libromayor['I8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['I8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['I8'] = 'DEBITO'
    sheet_libromayor['I8'].border = thin_border

    sheet_libromayor['J8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['J8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['J8'] = 'CREDITO'
    sheet_libromayor['J8'].border = thin_border

    sheet_libromayor['K8'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                             vertical="center")
    sheet_libromayor['K8'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['K8'] = 'SALDO'
    sheet_libromayor['K8'].border = thin_border

    saldo_actua = 0
    w = 0
    saldo_total = 0
    asientos = meses_asientos(data['detalle'])
    contador = 1
    for i in asientos:
        if contador != 1:
            saldo_actual_mes = 0
            saldo_actua = 0
        contador += 1
        sheet_libromayor['A' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                  horizontal="center",
                                                                                  vertical="center")
        sheet_libromayor['A' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                             color='000000')
        sheet_libromayor['A' + str(hoja_cuerpo)] = "Mes:"
        sheet_libromayor['A' + str(hoja_cuerpo)].border = thin_border

        sheet_libromayor['B' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                  horizontal="left",
                                                                                  vertical="center")
        sheet_libromayor['B' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                             color='000000')
        sheet_libromayor['B' + str(hoja_cuerpo)] = str(i)
        sheet_libromayor['B' + str(hoja_cuerpo)].border = thin_border
        hoja_cuerpo += 2
        for obj in asientos[i]:

            for l in obj:
                d = ""
                if len(data['documento']) != 0:
                    if w < len(data['documento']):
                        d = data['documento'][w]
                w += 1
                sheet_libromayor['A' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['A' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['A' + str(hoja_cuerpo)] = str(l.move_id.name)
                sheet_libromayor['A' + str(hoja_cuerpo)].border = thin_border


                sheet_libromayor['B' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['B' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['B' + str(hoja_cuerpo)] = str(l.ref or '')
                sheet_libromayor['B' + str(hoja_cuerpo)].border = thin_border

                sheet_libromayor['C' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['C' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['C' + str(hoja_cuerpo)] = str(l.move_id.date)
                sheet_libromayor['C' + str(hoja_cuerpo)].border = thin_border

                sheet_libromayor['D' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['D' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['D' + str(hoja_cuerpo)] = l.name
                sheet_libromayor['D' + str(hoja_cuerpo)].border = thin_border

                sheet_libromayor['E' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['E' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['E' + str(hoja_cuerpo)] = l.partner_id.name
                sheet_libromayor['E' + str(hoja_cuerpo)].border = thin_border

                sheet_libromayor['F' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['F' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['F' + str(hoja_cuerpo)] = l.jornada.name
                sheet_libromayor['F' + str(hoja_cuerpo)].border = thin_border
                sheet_libromayor['G' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['G' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['G' + str(hoja_cuerpo)] = l.seccion.name
                sheet_libromayor['G' + str(hoja_cuerpo)].border = thin_border
                sheet_libromayor['H' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="left",
                                                                                          vertical="center")
                sheet_libromayor['H' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['H' + str(hoja_cuerpo)] = l.curso.name
                sheet_libromayor['H' + str(hoja_cuerpo)].border = thin_border

                sheet_libromayor['I' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="right",
                                                                                          vertical="center")
                sheet_libromayor['I' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['I' + str(hoja_cuerpo)].number_format = '0.00'
                sheet_libromayor['I' + str(hoja_cuerpo)] = l.debit
                sheet_libromayor['I' + str(hoja_cuerpo)].border = thin_border

                sheet_libromayor['J' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="right",
                                                                                          vertical="center")
                sheet_libromayor['J' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                  color='000000')
                sheet_libromayor['J' + str(hoja_cuerpo)].number_format = '0.00'
                sheet_libromayor['J' + str(hoja_cuerpo)] = l.tax_amount
                sheet_libromayor['J' + str(hoja_cuerpo)].border = thin_border
                #_logger.info(int(data['cuenta'][0]))
                if int(data['cuenta'][0]) == 1:
                    saldo_actua += l.debit - l.tax_amount
                else:
                    saldo_actua += l.tax_amount - l.debit

                sheet_libromayor['K' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                          horizontal="right",
                                                                                          vertical="center")
                sheet_libromayor['K' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                                     color='000000')
                sheet_libromayor['K' + str(hoja_cuerpo)].number_format = '0.00'
                sheet_libromayor['K' + str(hoja_cuerpo)] = saldo_actua
                sheet_libromayor['K' + str(hoja_cuerpo)].border = thin_border
                hoja_cuerpo += 1

        saldo_actual_mes = saldo_actua
        sheet_libromayor['I' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                  horizontal="center",
                                                                                  vertical="center")
        sheet_libromayor['I' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=True, size=10,
                                                                             name="calibri", color='000000')
        sheet_libromayor['I' + str(hoja_cuerpo)] = "Saldo Mensual " + str(i) + " Total:"
        sheet_libromayor['I' + str(hoja_cuerpo)].border = thin_border
        sheet_libromayor['J' + str(hoja_cuerpo)].border = thin_border
        sheet_libromayor.merge_cells('I' + str(hoja_cuerpo) + ':J' + str(hoja_cuerpo))

        sheet_libromayor['K' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                                  horizontal="right",
                                                                                  vertical="center")
        sheet_libromayor['K' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10,
                                                                             name="calibri", color='000000')
        sheet_libromayor['K' + str(hoja_cuerpo)].number_format = '0.00'
        sheet_libromayor['K' + str(hoja_cuerpo)] = saldo_actual_mes
        sheet_libromayor['K' + str(hoja_cuerpo)].border = thin_border
        hoja_cuerpo += 2
        saldo_total += saldo_actual_mes
    sheet_libromayor['I' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                              horizontal="center",
                                                                              vertical="center")
    sheet_libromayor['I' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=True, size=10,
                                                                         name="calibri", color='000000')
    sheet_libromayor['I' + str(hoja_cuerpo)] = "Saldo Total: "
    sheet_libromayor['I' + str(hoja_cuerpo)].border = thin_border
    sheet_libromayor['J' + str(hoja_cuerpo)].border = thin_border
    sheet_libromayor.merge_cells('I' + str(hoja_cuerpo) + ':J' + str(hoja_cuerpo))

    sheet_libromayor['K' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                              horizontal="right",
                                                                              vertical="center")
    sheet_libromayor['K' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10,
                                                                         name="calibri", color='000000')
    sheet_libromayor['K' + str(hoja_cuerpo)].number_format = '0.00'
    sheet_libromayor['K' + str(hoja_cuerpo)] = saldo_total
    sheet_libromayor['K' + str(hoja_cuerpo)].border = thin_border
    hoja_cuerpo = hoja_cuerpo + 2




    return {'encabezado':hoja,'intro':hojas,'cuerpo':hoja_cuerpo,'total':saldo_total}

def totalizado_cuentas(sheet_libromayor,hoja_cuerpo,alignment_title,fuente_cabecera,saldo_total):
    hoja_cuerpo = hoja_cuerpo - 5
    sheet_libromayor['I' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                              horizontal="center",
                                                                              vertical="center")
    sheet_libromayor['I' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=True, size=10,
                                                                         name="calibri", color='000000')
    sheet_libromayor['I' + str(hoja_cuerpo)] = "Saldo Total Final : "
    sheet_libromayor['I' + str(hoja_cuerpo)].border = thin_border
    sheet_libromayor['J' + str(hoja_cuerpo)].border = thin_border
    sheet_libromayor.merge_cells('I' + str(hoja_cuerpo) + ':J' + str(hoja_cuerpo))

    sheet_libromayor['K' + str(hoja_cuerpo)].alignment = alignment_title.copy(wrapText=False,
                                                                              horizontal="right",
                                                                              vertical="center")
    sheet_libromayor['K' + str(hoja_cuerpo)].font = fuente_cabecera.copy(bold=False, size=10,
                                                                         name="calibri", color='000000')
    sheet_libromayor['K' + str(hoja_cuerpo)].number_format = '0.00'
    sheet_libromayor['K' + str(hoja_cuerpo)] = saldo_total
    sheet_libromayor['K' + str(hoja_cuerpo)].border = thin_border


def meses_asientos(data):
    dic = {}
    actual = ""
    datos = []
    i = 0

    for l in data:

        entered_date = datetime.strptime(l.date, '%Y-%m-%d')
        entered_date = entered_date.date()
        mes = mes_actual(entered_date.strftime("%B"))
        #_logger.info(str(len(data)) + " " + str(i) )
        if actual == "" or actual == mes:
            datos.append(l)
        else:
            dic[actual] = datos
            datos = []
            datos.append(l)
        actual = mes
        if i < len(data):
            dic[actual] = datos
        i += 1

    return dic




def size_celdas(valor,sheet_libromayor):
            sheet_libromayor.column_dimensions['A'].width = 14.45
            sheet_libromayor.column_dimensions['C'].width = 9.64
            sheet_libromayor.column_dimensions['B'].width = 14.00
            sheet_libromayor.column_dimensions['D'].width = 25.00
            sheet_libromayor.column_dimensions['E'].width = 30.20
            sheet_libromayor.column_dimensions['F'].width = 12.36
            sheet_libromayor.column_dimensions['G'].width = 12.91
            sheet_libromayor.column_dimensions['H'].width = 12.55
            sheet_libromayor.column_dimensions["I"].width = 11.09
            sheet_libromayor.column_dimensions['J'].width = 11.09
            sheet_libromayor.column_dimensions['K'].width = 10.82