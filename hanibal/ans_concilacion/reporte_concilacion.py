#-*- coding: utf-8 -*-
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import openpyxl.worksheet
import openpyxl
import datetime
from datetime import datetime
import logging
from openerp.osv import fields
_logger = logging.getLogger(__name__)
thin_border = Border(bottom=Side(style='thin'))
thin_border_left = Border(left=Side(style='thin'))
thin_border_top = Border(top=Side(style='thin'))
thin_border_right = Border(right=Side(style='thin'))
thin_border_left_top = Border(left=Side(style='thin'),top=Side(style='thin'))
thin_border_top_right = Border(top=Side(style='thin'),right=Side(style='thin'))
thin_border_top_right_buttom = Border(top=Side(style='thin'),right=Side(style='thin'),bottom=Side(style='thin'))
thin_border_buttonm_right = Border(right=Side(style='thin'),bottom=Side(style='thin'))
thin_border_buttonm_left = Border(left=Side(style='thin'),bottom=Side(style='thin'))
thin_border_buttonm_left_bottom = Border(left=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'))


def ajustes_hoja(sheet, flag, celda, value, value2):
    if (flag == 0):
        sheet.column_dimensions[celda].width = value2
    if (flag == 1):
        sheet.row_dimensions[int(celda)].height = value


def crear_hoja_libro(wb,title,flag,hojas):
    """if (flag == 0):
        sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
        # sheet.print_options.scale = 100
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        # sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidht = True
        # sheet.print_options.horizontalCentered = True
    if (flag == 1):
        #sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
        # sheet.print_options.scale = 100
        # sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = True
        sheet.page_margins.left = 0.1
        sheet.page_margins.right = 0.1
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        # sheet.print_options.horizontalCentered = True
    sheet.title = title """
    if(flag == 0):
        sheet = wb.active
	sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = hojas
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_margins.left = 0.6
        sheet.page_margins.right = 0.6
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5
    if(flag==1):
        sheet = wb.create_sheet()
	sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = hojas
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_margins.left = 0.6
        sheet.page_margins.right = 0.6
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

    sheet.title=title
    return sheet

def crear_reporte_excell():
    libro = openpyxl.Workbook(encoding="UTF-8", data_only=True)
    sheet_libromayor = crear_hoja_libro(libro, 'Reporte Concilacion Bancaria', 0, 0)
    sheet_view = openpyxl.worksheet.SheetView()
    sheet_view.zoomScale = "85"
    sheet_view.zoomScaleNormal = "70"
    sheet_libromayor.sheet_view = sheet_view
    sheet_libromayor.zoomScale = "70"
    ajustes_hoja(sheet_libromayor, 0, 'A', 20.00, 5.00)
    alignment_title = Alignment(horizontal='center', vertical='center')
    fuente_cabecera = Font(bold=True, size=15, name='calibri')

    all_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    dic = {'fuente':fuente_cabecera,'border':all_border,'alineacion':alignment_title,'sheet':sheet_libromayor,'libro':libro}

    return dic

def mes_actual(date):
    meses = {'January':'Enero','February':'Febrero','March':'Marzo','April':'Abril','May':'Mayo','June':'Junio'
        ,'July':'Julio','August':'Agosto','September':'Septiembe','October':'Octubre','November':'Noviembre','December':'Diciembre'}
    return meses[date]

def crear_encabezado(data,sheet_libromayor,alignment_title,fuente_cabecera):
    sheet_libromayor['A1'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['A1'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['A1'] = 'Compañia'


    sheet_libromayor['B1'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['B1'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B1'] = data['plan_cuenta'].name
    sheet_libromayor['B1'].border = thin_border_top


    sheet_libromayor['A1'].border = thin_border_left_top
    sheet_libromayor['A2'].border = thin_border_left
    sheet_libromayor['A3'].border = thin_border_left
    sheet_libromayor['A4'].border = thin_border_left
    sheet_libromayor['A4'].border = thin_border_left
    sheet_libromayor['A5'].border = thin_border_buttonm_left

    sheet_libromayor['G1'].border = thin_border_top_right
    sheet_libromayor['G2'].border = thin_border_right
    sheet_libromayor['G3'].border = thin_border_right
    sheet_libromayor['G4'].border = thin_border_right
    sheet_libromayor['G4'].border = thin_border_right
    sheet_libromayor['G5'].border = thin_border_buttonm_right

    sheet_libromayor['B5'].border = thin_border
    sheet_libromayor['C5'].border = thin_border
    sheet_libromayor['D5'].border = thin_border
    sheet_libromayor['E5'].border = thin_border
    sheet_libromayor['F5'].border = thin_border
    #sheet_libromayor['G5'].border = thin_border
    #sheet_libromayor['H5'].border = thin_border
    #sheet_libromayor['I5'].border = thin_border
    #sheet_libromayor['J5'].border = thin_border

    sheet_libromayor['B1'].border = thin_border_top
    sheet_libromayor['C1'].border = thin_border_top
    sheet_libromayor['D1'].border = thin_border_top
    sheet_libromayor['E1'].border = thin_border_top
    sheet_libromayor['F1'].border = thin_border_top
    #sheet_libromayor['G1'].border = thin_border_top
    #sheet_libromayor['H1'].border = thin_border_top_right
    #sheet_libromayor['I1'].border = thin_border_top
    #sheet_libromayor['J1'].border = thin_border_top_right
    #sheet_libromayor['K1'].border =

    sheet_libromayor['F1'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['F1'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['F1'] = 'Usuario:'

    #sheet_libromayor['H1'].border = thin_border_top
    sheet_libromayor['C1'].border = thin_border_top
    sheet_libromayor['D1'].border = thin_border_top
    sheet_libromayor['E1'].border = thin_border_top
    sheet_libromayor['F1'].border = thin_border_top
    #sheet_libromayor['G1'].border = thin_border_top
    sheet_libromayor['I1'].border = thin_border_top
    #sheet_libromayor['J1'].border = thin_border_top
    #sheet_libromayor['K1'].border = thin_border_top
    #sheet_libromayor.merge_cells('G1:H1')



    sheet_libromayor['G1'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['G1'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['G1'] = data['usuario']

    sheet_libromayor['A2'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['A2'].font = fuente_cabecera.copy(bold=True, size=20, name="calibri", color='000000')
    sheet_libromayor['A2'] = 'Concilacion Bancaria '
    sheet_libromayor.row_dimensions[2].height = float(31.50)

    sheet_libromayor.merge_cells('A2:G2')

    sheet_libromayor['D3'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['D3'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['D3'] = 'Desde: '

    sheet_libromayor['E3'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['E3'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['E3'] = data['fecha_inicial']

    sheet_libromayor['F3'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['F3'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['F3'] = 'Hasta: '
	
    sheet_libromayor['G3'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['G3'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['G3'] = data['fecha_final']
    hoja = 5
        
    sheet_libromayor['C5'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['C5'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['C5'] = 'Periodo: '
	
    sheet_libromayor['D5'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['D5'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    
    sheet_libromayor['D5'] = data['mes']
    
    sheet_libromayor['E5'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['E5'].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    
    sheet_libromayor['E5'] = data['ano']

    sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                       vertical="center")
    sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                  color='000000')
    sheet_libromayor['A' + str(hoja)] = 'Estado:'

    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                       vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                  color='000000')
    sheet_libromayor['B' + str(hoja)] = data['estado']

    sheet_libromayor['F5'].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                         vertical="center")
    sheet_libromayor['F5'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['F5'] = 'Fecha Emision: '

    sheet_libromayor['G5'].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                            vertical="center")
    sheet_libromayor['G5'].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['G5'] = fields.datetime.now()
    #sheet_libromayor.merge_cells('G5:H5')

    size_celdas(20.00, sheet_libromayor)



def encabezado_detalle(sheet_libromayor,alignment_title,fuente_cabecera,hoja):
    sheet_libromayor.row_dimensions[hoja].height = float(26.50)
    sheet_libromayor['A'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['A'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['A'+str(hoja)] = 'DOCUMENTOS'
    sheet_libromayor['A'+str(hoja)].border = thin_border_buttonm_left_bottom

    sheet_libromayor['B'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['B'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['B'+str(hoja)] = 'NUMERO'
    sheet_libromayor['B'+str(hoja)].border = thin_border

    sheet_libromayor['D'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['D'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['D'+str(hoja)] = 'VALOR'
    sheet_libromayor['D'+str(hoja)].border = thin_border


    sheet_libromayor['C'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['C'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['C'+str(hoja)] = 'FECHA'
    sheet_libromayor['C'+str(hoja)].border = thin_border


    sheet_libromayor['E'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                          vertical="center")
    sheet_libromayor['E'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['E'+str(hoja)] = 'DETALLE'
    sheet_libromayor['E'+str(hoja)].border = thin_border

    sheet_libromayor['F'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['F'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['F'+str(hoja)] = 'PROVEEDOR'
    sheet_libromayor['F'+str(hoja)].border = thin_border

    sheet_libromayor['G'+str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['G'+str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri", color='000000')
    sheet_libromayor['G'+str(hoja)] = 'MOVIMIENTO \nCONCILIADO'
    sheet_libromayor['G'+str(hoja)].border = thin_border_top_right_buttom


def detalle_saldos(dic,sheet_libromayor,alignment_title,fuente_cabecera):
    # ENCABEZADO DEL REPORTE CON VALORES
    hoja = 9
    sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                                       vertical="center")
    sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                  color='000000')
    sheet_libromayor['A' + str(hoja)] = 'Banco:'

    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                       vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                  color='000000')
    sheet_libromayor['B' + str(hoja)] = dic['Banco']

    hoja += 1

    sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="center",
                                                                       vertical="center")
    sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                  color='000000')
    sheet_libromayor['A' + str(hoja)] = 'Cuenta:'

    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                       vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                  color='000000')
    sheet_libromayor['B' + str(hoja)] = dic['Cuenta']


    hoja += 3
    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                            vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B' + str(hoja)] = 'Banco Cuenta Saldo'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                            vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F' + str(hoja)] = dic['saldo']
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'
    sheet_libromayor.merge_cells('B'+str(hoja)+':D'+str(hoja))
    hoja += 1

    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B' + str(hoja)] = '(+) Depositos no registrados'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                             vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F' + str(hoja)] = dic['DEP']
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'
    sheet_libromayor.merge_cells('B'+str(hoja)+':D'+str(hoja))
    hoja += 1

    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B' + str(hoja)] = '(-) Notas de Creditos'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                             vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F' + str(hoja)] = dic['NC']
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'
    sheet_libromayor.merge_cells('B' + str(hoja) + ':D' + str(hoja))
    hoja += 1
    sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['C' + str(hoja)] = 'S U B T O T A L'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                             vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    diferencia = (dic['saldo'] + dic['DEP']) - dic['NC']
    sheet_libromayor['F' + str(hoja)] = diferencia
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'
    #sheet_libromayor['E' + str(hoja)].border = thin_border_top
    sheet_libromayor['F' + str(hoja)].border = thin_border_top

    sheet_libromayor.merge_cells('C' + str(hoja) + ':D' + str(hoja))
    hoja += 1

    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B' + str(hoja)] = '(-) Cheques no registrados'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                             vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F' + str(hoja)] = dic['CHE']
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'
    sheet_libromayor.merge_cells('B' + str(hoja) + ':D' + str(hoja))
    hoja += 1

    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B' + str(hoja)] = '(-) Pago a Proveedores no registrados'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                             vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F' + str(hoja)] = dic['ORD']
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'
    sheet_libromayor.merge_cells('B' + str(hoja) + ':D' + str(hoja))
    hoja += 1
    sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['B' + str(hoja)] = '(-) N/Debitos no registrados'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                             vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F' + str(hoja)] = dic['ND']
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'
    sheet_libromayor.merge_cells('B' + str(hoja) + ':D' + str(hoja))
    hoja += 1
    sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                             vertical="center")
    sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['C' + str(hoja)] = 'SALDO CONCILIADO'

    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                             vertical="center")
    sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri", color='000000')
    sheet_libromayor['F' + str(hoja)] = diferencia - (dic['CHE'] + dic['ND'] + dic['ORD'])
    sheet_libromayor['F' + str(hoja)].border = thin_border_top
    #sheet_libromayor['E' + str(hoja)].border = thin_border_top
    sheet_libromayor['F' + str(hoja)].number_format = '#,##0.00'

    sheet_libromayor.merge_cells('C' + str(hoja) + ':D' + str(hoja))

def cuerpo_detalle(dic,sheet_libromayor,alignment_title,fuente_cabecera):
    hoja = 24
    if dic['d_che_c'] or dic['d_nd_c'] or dic['d_nc_c'] or dic['d_ord_c'] or dic['d_dep_c']:
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "CONCILIADOS"
        sheet_libromayor['A' + str(hoja)].border = thin_border
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        hoja += 1
    if dic['d_che_c']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "CHEQUES"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja)) #''
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_che_c']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle or ''
	    
	    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=True, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
	    sheet_libromayor.row_dimensions[hoja].height = float(26.50)
            sheet_libromayor['F' + str(hoja)] = l.move_id.company_char

            sheet_libromayor['G' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
	    #sheet_libromayor['F' + str(hoja)].style.alignment.wrap_text=True
            sheet_libromayor['G' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
	    documento = ""
	    for i in l.move_id:
		documento += i.name + ""
            sheet_libromayor['G' + str(hoja)] = documento or ''
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
	#sheet_libromayor['D' + str(hoja)].alignment.wrap_text = True
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2


    if dic['d_nd_c']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "NOTAS DE DEBITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_nd_c']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
	    
	    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=True, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
	    sheet_libromayor.row_dimensions[hoja].height = float(26.50)
            sheet_libromayor['F' + str(hoja)] = l.move_id.company_char

            sheet_libromayor['G' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['G' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            documento = ""
	    for i in l.move_id:
		documento += i.name + ""
            sheet_libromayor['G' + str(hoja)] = documento or ''
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_nc_c']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "NOTAS DE CREDITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_nc_c']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
	    
	    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=True, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
	    sheet_libromayor.row_dimensions[hoja].height = float(26.50)
            sheet_libromayor['F' + str(hoja)] = l.move_id.company_char

            sheet_libromayor['G' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['G' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
	
            documento = ""
	    for i in l.move_id:
		documento += i.name + ""
            sheet_libromayor['G' + str(hoja)] = documento or ''
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_dep_c']:
        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "DEPOSITOS"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_dep_c']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
	    
	    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=True, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
	    sheet_libromayor.row_dimensions[hoja].height = float(26.50)
            sheet_libromayor['F' + str(hoja)] = l.move_id.company_char

            sheet_libromayor['G' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['G' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            documento = ""
	    for i in l.move_id:
		documento += i.name + ""
            sheet_libromayor['G' + str(hoja)] = documento or ''
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_ord_c']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "PAGO A PROVEEDORES"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_che_c']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
	    
  	    sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=True, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
	    sheet_libromayor.row_dimensions[hoja].height = float(26.50)
            sheet_libromayor['F' + str(hoja)] = l.move_id.company_char

            sheet_libromayor['G' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['G' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')

            documento = ""
	    for i in l.move_id:
		documento += i.name + ""
            sheet_libromayor['G' + str(hoja)] = documento or ''
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_che_e'] or dic['d_nd_e'] or dic['d_nc_e'] or dic['d_ord_e'] or dic['d_dep_e']:
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "PENDIENTES POR LA EMPRESA"
        sheet_libromayor['A' + str(hoja)].border = thin_border
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        hoja += 1


    if dic['d_che_e']:
        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "CHEQUES"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_che_e']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                           color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                       color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_nd_e']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "NOTAS DE DEBITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_nd_e']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_nc_e']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "NOTAS DE CREDITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_nc_e']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_dep_e']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "DEPOSITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_dep_e']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_ord_e']:
        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "PAGO A PROVEEDORES"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_ord_e']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.parametro_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.ref

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.valor
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.fecha

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.detalle
            hoja += 1
            suma += l.valor
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2

    if dic['d_che_b'] or dic['d_nd_b'] or dic['d_nc_b'] or dic['d_ord_b'] or dic['d_dep_b']:
        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",color='000000')
        sheet_libromayor['A' + str(hoja)] = "PENDIENTES POR EL BANCO"
        sheet_libromayor['A' + str(hoja)].border = thin_border
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1

    if dic['d_che_b']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "CHEQUES"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_che_b']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.journal_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.name

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.total_conciliar
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.date

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.ref or ''

            sheet_libromayor['F' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['F' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['F' + str(hoja)] = l.company_char or ''
            hoja += 1
            suma += l.amount
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2





    if dic['d_nd_b']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "NOTAS DE DEBITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_nd_b']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.journal_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.name

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                         color='000000')
            sheet_libromayor['D' + str(hoja)] = l.total_conciliar
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.date

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.ref or ''
            hoja += 1
            suma += l.amount
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2





    if dic['d_nc_b']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "NOTAS DE CREDITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_nc_b']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.journal_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.name

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                        color='000000')

            sheet_libromayor['D' + str(hoja)] = l.total_conciliar
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.date

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.ref or ''
            hoja += 1
            suma += l.amount
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2




    if dic['d_dep_b']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "DEPOSITO"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_dep_b']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.journal_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.name

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')

            sheet_libromayor['D' + str(hoja)] = l.total_conciliar
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'


            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.date

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.ref or ''
            hoja += 1
            suma += l.amount
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2





    if dic['d_ord_b']:

        sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['A' + str(hoja)] = "PAGO A PROVEEDORES"
        sheet_libromayor['A' + str(hoja)].border = thin_border_buttonm_left_bottom
        sheet_libromayor['B' + str(hoja)].border = thin_border
        sheet_libromayor['C' + str(hoja)].border = thin_border
        sheet_libromayor['D' + str(hoja)].border = thin_border
        sheet_libromayor['E' + str(hoja)].border = thin_border
        sheet_libromayor['F' + str(hoja)].border = thin_border
        sheet_libromayor['G' + str(hoja)].border = thin_border_top_right_buttom
	
        #sheet_libromayor['H' + str(hoja)].border = thin_border
        sheet_libromayor.merge_cells('A' + str(hoja) + ':G' + str(hoja))
        hoja += 1
        suma = 0
        encabezado_detalle(sheet_libromayor, alignment_title, fuente_cabecera, hoja)
        hoja += 1
        for l in dic['d_ord_b']:
            sheet_libromayor['A' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['A' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['A' + str(hoja)] = l.journal_id.tipo_mov

            sheet_libromayor['B' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['B' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['B' + str(hoja)] = l.name

            sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                               vertical="center")
            sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['D' + str(hoja)] = l.total_conciliar
            sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'

            sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['C' + str(hoja)] = l.date

            sheet_libromayor['E' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                               vertical="center")
            sheet_libromayor['E' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                          color='000000')
            sheet_libromayor['E' + str(hoja)] = l.ref or ''
            hoja += 1
            suma += l.amount
        sheet_libromayor['C' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="left",
                                                                           vertical="center")
        sheet_libromayor['C' + str(hoja)].font = fuente_cabecera.copy(bold=True, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['C' + str(hoja)] = "Total"

        sheet_libromayor['D' + str(hoja)].alignment = alignment_title.copy(wrapText=False, horizontal="right",
                                                                           vertical="center")
        sheet_libromayor['D' + str(hoja)].font = fuente_cabecera.copy(bold=False, size=10, name="calibri",
                                                                      color='000000')
        sheet_libromayor['D' + str(hoja)] = suma
        sheet_libromayor['D' + str(hoja)].number_format = '#,##0.00'
        sheet_libromayor['D' + str(hoja)].border = thin_border_top
        hoja += 2











def size_celdas(valor,sheet_libromayor):
            sheet_libromayor.column_dimensions['A'].width = 14.45
            sheet_libromayor.column_dimensions['C'].width = 12.50
            sheet_libromayor.column_dimensions['B'].width = 20.00
            sheet_libromayor.column_dimensions['D'].width = 12.50
            sheet_libromayor.column_dimensions['E'].width = 24.50
            sheet_libromayor.column_dimensions['F'].width = 27.50
            sheet_libromayor.column_dimensions['G'].width = 25.64
            #sheet_libromayor.column_dimensions['H'].width = 15.18
            sheet_libromayor.column_dimensions["I"].width = 14.09
            sheet_libromayor.column_dimensions['J'].width = 11.09
            sheet_libromayor.column_dimensions['K'].width = 10.82
